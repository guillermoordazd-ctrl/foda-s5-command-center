"""
FODA JOPNAV S5 - Versión 5.2.1 (Alineación Estabilizada)
Interfaz táctica militar con análisis de matrices FODA, plano cartesiano,
simulaciones estocásticas Monte Carlo y motor de inteligencia por IA (Ollama).
"""

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio
import requests
import json
import warnings
pio.defaults.mathjax = None
warnings.filterwarnings("ignore", category=DeprecationWarning)
from datetime import datetime, date, timedelta
import io
import numpy as np
import multiprocessing
import re
import time
import os
from PIL import Image
from typing import cast, Any

# =============================================================================
# 1. CONFIGURACIÓN DE PÁGINA (DEBE SER EL PRIMER COMANDO DE STREAMLIT)
# =============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
icon_path = os.path.join(BASE_DIR, "my_foda_s5.ico")
icon_img = None
if os.path.exists(icon_path):
    try:
        icon_img = Image.open(icon_path)
    except Exception:
        pass

st.set_page_config(
    page_title="FODA INTELIGENTE",
    page_icon=icon_img if icon_img is not None else "my_foda_s5.ico",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================================================================
# 2. MÓDULOS DEL SISTEMA LOCAL
# =============================================================================
try:
    from database import init_db, save_analysis, get_all_analyses, clear_all_analyses
except ImportError:
    def init_db(): pass
    def save_analysis(*args, **kwargs): pass
    def get_all_analyses(): return []
    def clear_all_analyses(): pass

try:
    from pdf_utils import crear_pdf_final
except ImportError:
    def crear_pdf_final(*args, **kwargs): return b""

try:
    from nlp_mapper import map_columns_to_foda
except ImportError:
    def map_columns_to_foda(df): return {}

from models import (
    sanitizar_dataframe, calcular_scores, calcular_mefi_mefe, clasificacion,
    alertas, decidir, simular_periodos, simulacion_montecarlo,
    probabilidad_exito, nivel_riesgo, optimizar
)
from charts import (
    grafico_evolucion, grafico_simulacion, radar_estrategico,
    barras_balance, grafico_posicionamiento, crear_diagrama_ishikawa
)

# Caching helper for PDF generation to prevent CPU latency during tab navigation
@st.cache_data
def crear_pdf_final_cached(foda_dict, texto_ia, x, y, perfil="Analista", tipo_ana="General", fecha_i=None, fecha_t=None, prob_exito=50.0, bitacora=None, acciones=None, votos=None, total_votantes=10, came_estrategias=None, stress_int=0.0, stress_ext=0.0):
    return crear_pdf_final(
        foda_dict=foda_dict,
        texto_ia=texto_ia,
        x=x,
        y=y,
        perfil=perfil,
        tipo_ana=tipo_ana,
        fecha_i=fecha_i,
        fecha_t=fecha_t,
        prob_exito=prob_exito,
        bitacora=bitacora,
        acciones=acciones,
        votos=votos,
        total_votantes=total_votantes,
        came_estrategias=came_estrategias,
        stress_int=stress_int,
        stress_ext=stress_ext
    )

def apply_tactical_style():
    """
    Carga los estilos CSS desde el archivo styles.css e inyecta la tipografía Orbitron/Share Tech Mono.
    """
    try:
        css_path = os.path.join(BASE_DIR, "styles.css")
        with open(css_path, "r", encoding="utf-8") as f:
            css = f.read()
        st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)
    except Exception as e:
        st.warning(f"No se pudo cargar el archivo CSS de estilos externos: {str(e)}")

apply_tactical_style()

# =============================================================================
# 2. MOTOR DE IA Y FORMATEADORES (STREAMING)
# =============================================================================

def format_prompt_safe(template, **kwargs):
    result = template
    for key, val in kwargs.items():
        placeholder = "{" + key + "}"
        result = result.replace(placeholder, str(val))
    return result

def ejecutar_motor_ia_stream(perfil, tipo_ana, foda_data, fecha_i, fecha_t, x, y):
    """Genera el análisis militar/estratégico llamando a Ollama en modo stream (generador)."""
    prompt_tpl = st.session_state.get('prompt_foda', '')
    system_prompt = format_prompt_safe(
        prompt_tpl,
        perfil=perfil,
        tipo_ana=tipo_ana,
        fecha_i=fecha_i,
        fecha_t=fecha_t,
        x=x,
        y=y
    )
    user_payload = f"DATOS DE ENTRADA FODA: {json.dumps(foda_data, indent=2)}\n\nDIRECTIVA CRÍTICA DE IDIOMA: Responde única y exclusivamente en ESPAÑOL CASTELLANO. Queda terminantemente prohibido utilizar palabras en inglés, términos técnicos sin traducir o mezclar idiomas. Todo el contenido debe ser redactado en español puro."
    try:
        num_threads = max(1, multiprocessing.cpu_count() // 2)
        resp = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": "llama3", 
                "prompt": system_prompt + "\n\n" + user_payload, 
                "system": system_prompt, 
                "stream": True,
                "options": {
                    "num_thread": num_threads
                }
            },
            stream=True,
            timeout=90
        )
        for line in resp.iter_lines():
            if line:
                chunk = json.loads(line.decode('utf-8'))
                response_text = chunk.get('response', '')
                yield response_text
    except Exception as e:
        yield f"[ERROR DEL SISTEMA] No se pudo establecer enlace con Ollama en localhost:11434.\nDetalles: {str(e)}\n\n[RECOMENDACIÓN] Verifique que el servicio de Ollama esté activo con el modelo 'llama3' cargado."

def ejecutar_came_ia_stream(estrategia_tipo, foda_data):
    """Genera las recomendaciones tácticas CAME llamando a Ollama en modo stream."""
    detalles = {
        "A. Estrategia Ofensiva (F + O)": {
            "objetivo": "Maximizar fuerzas para aprovechar y capturar los beneficios y oportunidades del entorno.",
            "logica": "Se implementa cuando la organización posee una ventaja competitiva clara y el entorno presenta condiciones favorables. Consiste en enfocarse en eso que somos buenos y en lo que no lo son los demás, para poder aprovechar las oportunidades a las que la competencia no podría llegar."
        },
        "B. Estrategia de Supervivencia (D + A)": {
            "objetivo": "Minimizar daños, corregir debilidades internas y resistir la hostilidad externa del entorno.",
            "logica": "Queremos sobrevivir en el mercado, por lo que debemos pensar en cómo mitigar o eliminar las debilidades para dar cara a las amenazas o que al menos, las amenazas no terminen por desaparecernos."
        },
        "C. Estrategia Defensiva (F + A)": {
            "objetivo": "Usar capacidades y fortalezas consolidadas como escudo táctico para hacer frente a amenazas.",
            "logica": "Consiste en la explotación de las fortalezas para hacer frente a las amenazas. Su objetivo consiste en mantener la posición conseguida en el mercado."
        },
        "D. Estrategia de Reorientación (D + O)": {
            "objetivo": "Corregir brechas operativas y debilidades internas para capturar oportunidades del entorno.",
            "logica": "Busca detectar en dónde se ha estado fallando a nivel interno (debilidades) para empezar a aprovechar oportunidades que no podríamos conseguir sin antes tratar dichas fallas."
        }
    }
    
    info = detalles.get(estrategia_tipo, {
        "objetivo": "Formular recomendaciones estratégicas.",
        "logica": "Cruzar variables internas y externas."
    })
    
    prompt_tpl = st.session_state.get('prompt_came', '')
    system_prompt = format_prompt_safe(
        prompt_tpl,
        estrategia_tipo=estrategia_tipo,
        objetivo=info['objetivo'],
        logica=info['logica']
    )
    
    user_payload = f"FACTORES FODA ACTIVOS SUMINISTRADOS: {json.dumps(foda_data, indent=2)}\n\nDIRECTIVA CRÍTICA DE IDIOMA: Responde única y exclusivamente en ESPAÑOL CASTELLANO. Queda terminantemente prohibido utilizar palabras en inglés, términos técnicos sin traducir o mezclar idiomas. Todo el contenido debe ser redactado en español puro."
    try:
        num_threads = max(1, multiprocessing.cpu_count() // 2)
        resp = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": "llama3", 
                "prompt": system_prompt + "\n\n" + user_payload, 
                "system": system_prompt, 
                "stream": True,
                "options": {
                    "num_thread": num_threads
                }
            },
            stream=True,
            timeout=90
        )
        for line in resp.iter_lines():
            if line:
                chunk = json.loads(line.decode('utf-8'))
                response_text = chunk.get('response', '')
                yield response_text
    except Exception as e:
        yield f"[ERROR DEL SISTEMA] No se pudo establecer enlace con Ollama en localhost:11434.\nDetalles: {str(e)}\n\n[RECOMENDACIÓN] Verifique que el servicio de Ollama esté activo con el modelo 'llama3' cargado."

def quitar_emojis(texto):
    if not texto:
        return ""
    # Remover SMP emojis (plano de emojis regular, ej. circulos de colores, manos, etc.)
    clean_text = re.sub(r'[\U00010000-\U0010ffff]', '', texto)
    # Remover dingbats, formas geométricas y flechas especiales
    clean_text = re.sub(r'[\u25a0-\u27bf]', '', clean_text)
    # Remover caracteres de la tabla de símbolos misceláneos
    clean_text = re.sub(r'[\u2600-\u26ff]', '', clean_text)
    return clean_text

def sanitizar_emojis_texto_ia(texto):
    if not texto:
        return ""
    lines = texto.split('\n')
    new_lines = []
    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            new_lines.append(line)
            continue
        
        # Determinar si es un encabezado principal de sección
        is_header = (
            any(h in line_stripped for h in ["SITUACIÓN GENERAL", "AMENAZAS", "RIESGOS", "LÍNEAS DE ACCIÓN"]) or
            any(e in line_stripped for e in ["🔵", "🟡", "🔴", "🟢"])
        )
        
        if is_header:
            # Detectar si la línea comienza con un emoji círculo y conservarlo
            emoji_prefix = ""
            for e in ["🔵", "🟡", "🔴", "🟢"]:
                if line_stripped.startswith(e):
                    emoji_prefix = e + " "
                    line_stripped = line_stripped[len(e):].strip()
                    break
            # Limpiar el resto de la línea de otros emojis residuales
            line_cleaned = emoji_prefix + quitar_emojis(line_stripped)
            new_lines.append(line_cleaned)
        else:
            # Línea normal de cuerpo o lista: quitar todos los emojis
            new_lines.append(quitar_emojis(line))
            
    return '\n'.join(new_lines)

def normalizar_espaciado(texto):
    if not texto:
        return ""
    lines = [line.strip() for line in texto.split('\n')]
    new_lines = []
    for line in lines:
        if not line:
            continue
        # Determinar si es un encabezado principal (por palabra clave, emoji o numeración CAME)
        is_header = (
            any(h in line for h in ["SITUACIÓN GENERAL", "AMENAZAS", "RIESGOS", "LÍNEAS DE ACCIÓN"]) or
            any(e in line for e in ["🔵", "🟡", "🔴", "🟢"]) or
            re.match(r'^(1\.[^0-9]|2\.[^0-9]|3\.\s*(?:l[ií]neas|recom|accion|dirección|postura))', line, re.IGNORECASE)
        )
        if is_header:
            if new_lines:
                new_lines.append("") # Agregar una línea vacía antes del nuevo encabezado (interlineado doble)
            new_lines.append(line)
        else:
            new_lines.append(line)
    return "\n".join(new_lines)

def format_came_estrategia(text):
    if not text:
        return ""
    if "[ERROR DEL SISTEMA]" in text:
        return f'<div class="report-box-fallback" style="color: #ff0055 !important; font-size: 18px !important; text-align: justify; margin-left: 20px;">{text}</div>'
    
    # 1. Normalizar espaciado
    text_normalized = normalizar_espaciado(text)
    
    # 2. Agrupar por secciones
    lines = [line.strip() for line in text_normalized.split('\n') if line.strip()]
    
    part1_header = ""
    part1_paragraphs = []
    part2_header = ""
    part2_paragraphs = []
    recs = []
    
    current_part = 0
    
    for line in lines:
        if (re.match(r'^1\.[^0-9]', line, re.IGNORECASE) or "1. Fortalezas" in line or "1. Debilidades" in line) and current_part < 1:
            current_part = 1
            match = re.match(r'^(1\.[^:]+:?)(.*)', line, re.IGNORECASE)
            if match:
                part1_header = match.group(1).strip()
                body_init = match.group(2).strip()
                if body_init:
                    part1_paragraphs.append(body_init)
            else:
                part1_header = line
            continue
        elif (re.match(r'^2\.[^0-9]', line, re.IGNORECASE) or "2. Oportunidades" in line or "2. Amenazas" in line) and current_part < 2:
            current_part = 2
            match = re.match(r'^(2\.[^:]+:?)(.*)', line, re.IGNORECASE)
            if match:
                part2_header = match.group(1).strip()
                body_init = match.group(2).strip()
                if body_init:
                    part2_paragraphs.append(body_init)
            else:
                part2_header = line
            continue
        elif re.match(r'^(3\.\s*(?:l[ií]neas|recom|accion|dirección|postura)|3\.,\s*4\.)', line, re.IGNORECASE) and current_part < 3:
            current_part = 3
            continue
            
        if current_part == 1:
            part1_paragraphs.append(line)
        elif current_part == 2:
            part2_paragraphs.append(line)
        elif current_part == 3 or current_part == 0:
            cleaned_rec = re.sub(r'^\s*(?:[\*\-\+]\s+|(?:\d+[\.\)\s]*)+)', '', line).strip()
            if cleaned_rec:
                recs.append(cleaned_rec)
                
    # Fallback si no se detecta la estructura esperada
    if not part1_paragraphs and not part2_paragraphs and not recs:
        fallback_paragraphs = [p.strip() for p in text_normalized.split('\n') if p.strip()]
        html_fallback = '<div style="background: rgba(2, 6, 23, 0.7); border: 1px solid rgba(0, 242, 255, 0.2); border-radius: 6px; padding: 22px; box-shadow: inset 0 0 15px rgba(0, 242, 255, 0.05);">'
        for idx, p in enumerate(fallback_paragraphs):
            indent_style = "text-indent: 25px;" if idx > 0 else ""
            html_fallback += f'<p style="font-family: \'Share Tech Mono\', monospace; font-size: 18px; text-align: justify; line-height: 1.35; color: #e2e8f0; margin: 0 0 10px 0; margin-left: 20px; {indent_style}">{p}</p>'
        html_fallback += '</div>'
        return "".join(line.strip() for line in html_fallback.split("\n"))
        
    html = '<div style="background: rgba(2, 6, 23, 0.7); border: 1px solid rgba(0, 242, 255, 0.2); border-radius: 6px; padding: 22px; box-shadow: inset 0 0 15px rgba(0, 242, 255, 0.05);">'
    
    if part1_paragraphs:
        header_display = part1_header if part1_header else "1. ANÁLISIS INTERNO"
        html += '<div style="margin-bottom: 20px; border-left: 3px solid #39FF14; padding-left: 15px;">'
        html += f'<div style="font-family: \'Orbitron\', sans-serif; font-size: 13px; font-weight: bold; color: #39FF14; text-shadow: 0 0 5px rgba(57,255,20,0.3); margin-bottom: 4px; text-transform: uppercase;">{header_display}</div>'
        for idx, p in enumerate(part1_paragraphs):
            indent_style = "text-indent: 25px;" if idx > 0 else ""
            html += f'<p style="font-family: \'Share Tech Mono\', monospace; font-size: 18px; text-align: justify; line-height: 1.35; color: #e2e8f0; margin: 0 0 10px 0; margin-left: 20px; {indent_style}">{p}</p>'
        html += '</div>'
        
    if part2_paragraphs:
        header_display = part2_header if part2_header else "2. ANÁLISIS EXTERNO"
        html += '<div style="margin-bottom: 25px; border-left: 3px solid #00f2fe; padding-left: 15px;">'
        html += f'<div style="font-family: \'Orbitron\', sans-serif; font-size: 13px; font-weight: bold; color: #00f2fe; text-shadow: 0 0 5px rgba(0,242,254,0.3); margin-bottom: 4px; text-transform: uppercase;">{header_display}</div>'
        for idx, p in enumerate(part2_paragraphs):
            indent_style = "text-indent: 25px;" if idx > 0 else ""
            html += f'<p style="font-family: \'Share Tech Mono\', monospace; font-size: 18px; text-align: justify; line-height: 1.35; color: #e2e8f0; margin: 0 0 10px 0; margin-left: 20px; {indent_style}">{p}</p>'
        html += '</div>'
        
    if recs:
        html += '<div style="border-top: 1px solid rgba(255,255,255,0.07); padding-top: 18px;">'
        html += '<div style="font-family: \'Orbitron\', sans-serif; font-size: 13px; font-weight: bold; color: #ffaa00; text-shadow: 0 0 5px rgba(255,170,0,0.3); margin-bottom: 12px; text-transform: uppercase; letter-spacing: 1px;">3. LÍNEAS DE ACCIÓN Y RECOMENDACIONES</div>'
        html += '<div style="display: flex; flex-direction: column; gap: 10px;">'
        for idx, rec in enumerate(recs):
            num_val = idx + 1
            html += f'<div style="display: flex; align-items: flex-start; font-family: \'Share Tech Mono\', monospace; font-size: 18px; color: #cbd5e1; line-height: 1.35; margin-left: 20px;">'
            html += f'<span style="font-family: \'Orbitron\', sans-serif; font-weight: bold; color: #ffaa00; text-shadow: 0 0 4px rgba(255,170,0,0.3); min-width: 25px; display: inline-block;">{num_val}.</span>'
            html += f'<span style="flex-grow: 1; text-align: justify; padding-left: 5px;">{rec}</span>'
            html += '</div>'
        html += '</div></div>'
        
    html += '</div>'
    return "".join(line.strip() for line in html.split("\n"))

def format_list_items(text, color_class):
    if not text:
        return ""
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    items = []
    for line in lines:
        cleaned_line = re.sub(r'^\s*(?:[\*\-\+]\s+|(?:\d+\.(?!\d)|\d+\))\s*)', '', line).strip()
        # Eliminar emojis indicadores en cualquier parte del ítem para evitar ruido visual
        cleaned_line = re.sub(r'[🔴🟡🟢🔵]', '', cleaned_line).strip()
        # Ignorar líneas que no contienen letras o números (por ejemplo, emojis huérfanos)
        if cleaned_line and any(c.isalnum() for c in cleaned_line):
            items.append(cleaned_line)
            
    if not items:
        return "<p style='font-size: 20px !important;'>Sin incidencias registradas.</p>"
        
    html = f'<ol class="{color_class}">'
    for item in items:
        html += f"<li>{item}</li>"
    html += "</ol>"
    return "".join(line.strip() for line in html.split("\n"))

def format_ia_report(text):
    if not text:
        return ""
    if "[ERROR DEL SISTEMA]" in text:
        return f'<div class="report-box-fallback" style="color: #ff0055 !important;">{text}</div>'
    
    headers = [
        ("SITUACIÓN GENERAL", [r"situaci[oó]n\s+general"]),
        ("AMENAZAS", [r"amenazas"]),
        ("RIESGOS", [r"riesgos"]),
        ("LÍNEAS DE ACCIÓN", [r"l[ií]neas\s+de\s+acci[oó]n", r"lineas\s+de\s+accion"])
    ]
    
    section_indices = []
    for title, regexes in headers:
        for regex in regexes:
            pattern = r"(?:^|\n)[^a-zA-Z0-9\n]*" + regex + r"[:#\*\-\s]*"
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                section_indices.append((match.start(), match.end(), title))
                break
                
    section_indices.sort(key=lambda x: x[0])
    
    sections = {
        "SITUACIÓN GENERAL": "",
        "AMENAZAS": "",
        "RIESGOS": "",
        "LÍNEAS DE ACCIÓN": ""
    }
    
    for i in range(len(section_indices)):
        start_idx, end_idx, title = section_indices[i]
        next_start = section_indices[i+1][0] if i + 1 < len(section_indices) else len(text)
        sections[title] = text[end_idx:next_start].strip()
        
    if not any(sections.values()):
        fallback_paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
        html_fallback = '<div class="report-container">'
        for idx, p in enumerate(fallback_paragraphs):
            indent_style = "text-indent: 25px;" if idx > 0 else ""
            html_fallback += f'<p style="font-family: \'Share Tech Mono\', monospace; font-size: 20px !important; line-height: 1.15 !important; text-align: justify; color: #e2e8f0; margin: 0 0 10px 0; margin-left: 25px; {indent_style}">{p}</p>'
        html_fallback += '</div>'
        return "".join(line.strip() for line in html_fallback.split("\n"))
        
    html_output = '<div class="report-container">'
    
    # 1. SITUACIÓN GENERAL
    sit_text = sections["SITUACIÓN GENERAL"]
    if sit_text.strip():
        sit_paragraphs = []
        for line in sit_text.split('\n'):
            line_strip = line.strip()
            if line_strip:
                cleaned_line = re.sub(r'^\s*(?:\d+[\.\-\s\)]+|[\*\-\+])\s*', '', line_strip).strip()
                cleaned_line = re.sub(r'[🔴🟡🟢🔵]', '', cleaned_line).strip()
                if cleaned_line:
                    sit_paragraphs.append(cleaned_line)
                    
        sit_html = ""
        for idx, p in enumerate(sit_paragraphs):
            indent_style = "text-indent: 25px;" if idx > 0 else ""
            sit_html += f'<p style="margin: 0 0 10px 0; text-align: justify; {indent_style}">{p}</p>'
        
        html_output += f"""
        <div class="report-section">
            <div class="report-title title-blue">🔵 SITUACIÓN GENERAL</div>
            <div class="report-body prose-body">{sit_html}</div>
        </div>
        """
    
    # 2. AMENAZAS
    if sections["AMENAZAS"].strip():
        html_output += f"""
        <div class="report-section">
            <div class="report-title title-yellow">🟡 AMENAZAS</div>
            <div class="report-body list-body">{format_list_items(sections["AMENAZAS"], "ol-yellow")}</div>
        </div>
        """
    
    # 3. RIESGOS
    if sections["RIESGOS"].strip():
        html_output += f"""
        <div class="report-section">
            <div class="report-title title-red">🔴 RIESGOS</div>
            <div class="report-body list-body">{format_list_items(sections["RIESGOS"], "ol-red")}</div>
        </div>
        """
    
    # 4. LÍNEAS DE ACCIÓN
    if sections["LÍNEAS DE ACCIÓN"].strip():
        html_output += f"""
        <div class="report-section">
            <div class="report-title title-green">🟢 LÍNEAS DE ACCIÓN</div>
            <div class="report-body list-body">{format_list_items(sections["LÍNEAS DE ACCIÓN"], "ol-green")}</div>
        </div>
        """
    
    html_output += '</div>'
    return "".join(line.strip() for line in html_output.split("\n"))

# =============================================================================
# 4. INICIALIZACIÓN DEL ESTADO DE SESIÓN
# =============================================================================
def reset_pdf_state():
    st.session_state.pdf_ready_informe = False
    st.session_state.pdf_ready_mando = False
    st.session_state.pop("pdf_bytes_informe", None)
    st.session_state.pop("pdf_bytes_mando", None)

def limpiar_reporte_ia():
    st.session_state.ultimo_resultado = ""
    reset_pdf_state()

def limpiar_analisis_memoria():
    st.session_state.ultimo_resultado = ""
    st.session_state.came_estrategias = {
        "A. Estrategia Ofensiva (F + O)": "",
        "B. Estrategia de Supervivencia (D + A)": "",
        "C. Estrategia Defensiva (F + A)": "",
        "D. Estrategia de Reorientación (D + O)": ""
    }
    reset_pdf_state()




def make_progress_cb(progress_bar):
    def cb(pct, text):
        progress_bar.progress(pct, text=text)
        time.sleep(0.08)  # Pequeña pausa táctica para ver la compilación fluir
    return cb


def safe_to_image(fig, format="png", width=800, height=600, scale=1.0):
    """
    Exporta una figura de Plotly a imagen en bytes.
    Si kaleido requiere Google Chrome/Chromium y no se encuentra instalado en el sistema,
    intenta descargarlo dinámicamente de forma silenciosa esquivando errores de certificados SSL
    y reintenta la exportación.
    """
    import ssl
    try:
        return fig.to_image(format=format, width=width, height=height, scale=scale)
    except Exception as e:
        err_msg = str(e).lower()
        if "chrome" in err_msg or "browser" in err_msg or "not found" in err_msg or "choreographer" in err_msg:
            try:
                import kaleido
                if hasattr(kaleido, "get_chrome_sync"):
                    # Evitar errores de verificación de certificados SSL al descargar Chromium en macOS
                    old_context = getattr(ssl, "_create_default_https_context", None)
                    if old_context is not None:
                        setattr(ssl, "_create_default_https_context", ssl._create_unverified_context)
                    try:
                        getattr(kaleido, "get_chrome_sync")()
                    finally:
                        if old_context is not None:
                            setattr(ssl, "_create_default_https_context", old_context)
                    
                    # Reintentar la exportación
                    return fig.to_image(format=format, width=width, height=height, scale=scale)
                else:
                    raise e
            except Exception as inner_e:
                raise RuntimeError(
                    f"No se pudo descargar el navegador automatizado requerido para imágenes en Excel. "
                    f"Por favor, conéctese a Internet, instale Google Chrome o ejecute 'plotly_get_chrome'. "
                    f"Detalle: {inner_e}"
                ) from e
        raise e


def crear_excel_estilizado(foda_data, x, y, prob_exito, perfil, tipo_ana, f_inicio, f_fin, bitacora, votos, total_votantes, came_estrategias, alert_level, ultimo_resultado, acciones):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.drawing.image import Image as OpenpyxlImage

    wb = Workbook()
    # Eliminar hoja por defecto
    default_sheet = cast(Worksheet, wb.active)
    wb.remove(default_sheet)

    # ----------------------------------------------------
    # Definición de Estilos
    # ----------------------------------------------------
    font_family = "Segoe UI"

    # Fuentes
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    font_bold = Font(name=font_family, size=11, bold=True, color="000000")
    font_bold_white = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_regular = Font(name=font_family, size=11, color="000000")
    font_italic = Font(name=font_family, size=10, italic=True, color="555555")

    # Rellenos (Fills)
    fill_navy = PatternFill(start_color="0B1329", end_color="0B1329", fill_type="solid")
    fill_light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Rellenos de categorías FODA (cabeceras oscuras y tonos suaves para celdas)
    fill_fortaleza_hdr = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Verde oscuro
    fill_debilidad_hdr = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid") # Rojo oscuro
    fill_oportunidad_hdr = PatternFill(start_color="006064", end_color="006064", fill_type="solid") # Teal oscuro
    fill_amenaza_hdr = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid") # Naranja oscuro

    fill_fortaleza_light = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") # Verde claro
    fill_debilidad_light = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid") # Rojo claro
    fill_oportunidad_light = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid") # Teal claro
    fill_amenaza_light = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid") # Naranja claro

    # Bordes
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thick_border_side = Side(border_style="medium", color="0B1329")
    double_border_side = Side(border_style="double", color="0B1329")

    border_all_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total_row = Border(top=thin_border_side, bottom=double_border_side)

    # Alineaciones
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def create_title_banner(ws, title_text, max_col):
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=max_col)
        for r in range(2, 4):
            for c in range(2, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_navy
        title_cell = ws.cell(row=2, column=2)
        title_cell.value = title_text
        title_cell.font = font_title
        title_cell.alignment = align_center

    # ----------------------------------------------------
    # Hojas: 1. Resumen Ejecutivo
    # ----------------------------------------------------
    ws_res = cast(Worksheet, wb.create_sheet(title="Resumen Ejecutivo"))
    cast(Any, ws_res.views.sheetView)[0].showGridLines = True
    create_title_banner(ws_res, "FODA JOPNAV S5 // INFORME ESTRATÉGICO EJECUTIVO", 6)

    metadata = [
        ("FECHA DE EMISIÓN", datetime.now().strftime("%Y-%m-%d")),
        ("PERFIL OPERATIVO", perfil),
        ("TIPO DE ANÁLISIS", tipo_ana),
        ("VENTANA TEMPORAL", f"{f_inicio} a {f_fin}"),
        ("VECTOR POSICIONAMIENTO", f"X: {x:.3f} (MEFI) // Y: {y:.3f} (MEFE)"),
        ("PROBABILIDAD DE ÉXITO MC", f"{prob_exito:.1f}%"),
        ("ESTADO DE ALERTA OPERATIVA", alert_level)
    ]

    row_idx = 5
    for key, val in metadata:
        c1 = ws_res.cell(row=row_idx, column=2, value=key)
        c1.font = font_bold
        c1.fill = fill_light_gray
        c1.alignment = align_left
        c1.border = border_all_thin

        c2 = ws_res.cell(row=row_idx, column=3, value=val)
        c2.font = font_regular
        c2.alignment = align_left
        c2.border = border_all_thin

        if key == "ESTADO DE ALERTA OPERATIVA":
            if val == "ROJO":
                c2.font = Font(name=font_family, size=11, bold=True, color="FF0055")
            elif val == "AMARILLO":
                c2.font = Font(name=font_family, size=11, bold=True, color="FFAA00")
            else:
                c2.font = Font(name=font_family, size=11, bold=True, color="2E7D32")
        row_idx += 1

    row_idx += 2
    ws_res.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
    sec_hdr = ws_res.cell(row=row_idx, column=2, value="INFORME DE INTELIGENCIA ESTRATÉGICA (IA)")
    sec_hdr.font = font_subtitle
    sec_hdr.fill = fill_navy
    sec_hdr.alignment = align_center
    for c in range(2, 7):
        ws_res.cell(row=row_idx, column=c).border = border_all_thin
    row_idx += 1

    # Parsear el reporte de IA
    report_sections = {
        "🔵 SITUACIÓN GENERAL": "",
        "🟡 AMENAZAS DETECTADAS": "",
        "🔴 RIESGOS CRÍTICOS": "",
        "🟢 LÍNEAS DE ACCIÓN RECOMENDADAS": ""
    }

    text = ultimo_resultado or ""
    headers_regex = [
        ("🔵 SITUACIÓN GENERAL", [r"situaci[oó]n\s+general"]),
        ("🟡 AMENAZAS DETECTADAS", [r"amenazas"]),
        ("🔴 RIESGOS CRÍTICOS", [r"riesgos"]),
        ("🟢 LÍNEAS DE ACCIÓN RECOMENDADAS", [r"l[ií]neas\s+de\s+acci[oó]n", r"lineas\s+de\s+accion"])
    ]

    section_indices = []
    for title, regexes in headers_regex:
        for regex in regexes:
            pattern = r"(?:^|\n)[^a-zA-Z0-9\n]*" + regex + r"[:#\*\-\s]*"
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                section_indices.append((match.start(), match.end(), title))
                break
    section_indices.sort(key=lambda x: x[0])

    if section_indices:
        for i in range(len(section_indices)):
            start_idx, end_idx, title = section_indices[i]
            next_start = section_indices[i+1][0] if i + 1 < len(section_indices) else len(text)
            report_sections[title] = text[end_idx:next_start].strip()
    else:
        report_sections["🔵 SITUACIÓN GENERAL"] = text

    for title, content in report_sections.items():
        if content.strip():
            # Encabezado sección IA
            ws_res.cell(row=row_idx, column=2, value=title).font = font_bold
            ws_res.cell(row=row_idx, column=2).alignment = align_left
            ws_res.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
            for col_c in range(2, 7):
                ws_res.cell(row=row_idx, column=col_c).border = Border(bottom=thin_border_side)
            row_idx += 1

            lines = [l.strip() for l in content.split('\n') if l.strip()]
            for line in lines:
                cleaned_line = re.sub(r'^\s*(?:[\*\-\+]\s+|(?:\d+[\.\)\s]*)+)', '', line).strip()
                cleaned_line = re.sub(r'[🔴🟡🟢🔵]', '', cleaned_line).strip()
                if not cleaned_line:
                    continue

                if title != "🔵 SITUACIÓN GENERAL":
                    # Lista con viñeta
                    ws_res.cell(row=row_idx, column=2, value="•").alignment = align_center
                    ws_res.cell(row=row_idx, column=2).font = font_bold
                    ws_res.cell(row=row_idx, column=2).border = border_all_thin

                    c_txt = ws_res.cell(row=row_idx, column=3, value=cleaned_line)
                    c_txt.font = font_regular
                    c_txt.alignment = align_wrap_left
                    ws_res.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=6)
                    for col_c in range(3, 7):
                        ws_res.cell(row=row_idx, column=col_c).border = border_all_thin
                    row_idx += 1
                else:
                    # Párrafo Situación General
                    ws_res.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx+1, end_column=6)
                    c_txt = ws_res.cell(row=row_idx, column=2, value=cleaned_line)
                    c_txt.font = font_regular
                    c_txt.alignment = align_wrap_left
                    for r_sub in range(row_idx, row_idx + 2):
                        for col_c in range(2, 7):
                            ws_res.cell(row=r_sub, column=col_c).border = border_all_thin
                    row_idx += 2
            row_idx += 1

    # ----------------------------------------------------
    # Hojas: 1.5 Diagnóstico Estratégico (Gráficas)
    # ----------------------------------------------------
    df_f = pd.DataFrame(foda_data.get("F", []))
    df_d = pd.DataFrame(foda_data.get("D", []))
    df_o = pd.DataFrame(foda_data.get("O", []))
    df_a = pd.DataFrame(foda_data.get("A", []))
    
    scores = calcular_mefi_mefe(df_f, df_d, df_o, df_a)
    estado_actual = clasificacion(scores)
    
    ws_diag = cast(Worksheet, wb.create_sheet(title="Diagnóstico Estratégico"))
    cast(Any, ws_diag.views.sheetView)[0].showGridLines = False
    create_title_banner(ws_diag, "DIAGNÓSTICO ESTRATÉGICO Y ANÁLISIS DE BALANCE", 18)
    
    # Configurar dimensiones de celdas para evitar amontonamiento y lograr alineación impecable
    for col_idx in range(2, 27):
        col_letter = get_column_letter(col_idx)
        ws_diag.column_dimensions[col_letter].width = 11
    for r_idx in range(5, 50):
        ws_diag.row_dimensions[r_idx].height = 20
        
    try:
        # Generar figuras
        fig_pos = grafico_posicionamiento(x, y)
        fig_radar = radar_estrategico(scores)
        fig_bar = barras_balance(scores)
        
        # Convertir a imágenes PNG optimizadas y ultra-nítidas (escala 1.3)
        img_pos_bytes = safe_to_image(fig_pos, format="png", width=500, height=380, scale=1.3)
        img_radar_bytes = safe_to_image(fig_radar, format="png", width=500, height=380, scale=1.3)
        img_bar_bytes = safe_to_image(fig_bar, format="png", width=1020, height=320, scale=1.3)
        
        # Envolver en openpyxl Image
        img_pos = OpenpyxlImage(io.BytesIO(img_pos_bytes))
        img_radar = OpenpyxlImage(io.BytesIO(img_radar_bytes))
        img_bar = OpenpyxlImage(io.BytesIO(img_bar_bytes))
        
        # Insertar en la hoja respetando márgenes y espaciado (B5, K5 y B26)
        ws_diag.add_image(img_pos, "B5")
        ws_diag.add_image(img_radar, "K5")
        ws_diag.add_image(img_bar, "B26")
    except Exception as e:
        ws_diag.cell(row=5, column=2, value=f"Error al generar gráficos: {e}").font = font_italic

    # ----------------------------------------------------
    # Hojas: 1.6 Análisis Causa-Efecto (Ishikawa)
    # ----------------------------------------------------
    ws_ish = cast(Worksheet, wb.create_sheet(title="Análisis Causa-Efecto"))
    cast(Any, ws_ish.views.sheetView)[0].showGridLines = False
    create_title_banner(ws_ish, "ISHIKAWA // DIAGRAMA DE CAUSA-EFECTO EN TIEMPO REAL", 18)
    
    # Configurar dimensiones de celdas para mantener consistencia visual
    for col_idx in range(2, 27):
        col_letter = get_column_letter(col_idx)
        ws_ish.column_dimensions[col_letter].width = 11
    for r_idx in range(5, 40):
        ws_ish.row_dimensions[r_idx].height = 20
        
    try:
        # Generar figura
        fig_ish = crear_diagrama_ishikawa(df_d, df_a, estado_actual)
        
        # Convertir a imagen PNG optimizada y ultra-nítida (escala 1.3)
        img_ish_bytes = safe_to_image(fig_ish, format="png", width=1020, height=450, scale=1.3)
        
        # Envolver e insertar
        img_ish = OpenpyxlImage(io.BytesIO(img_ish_bytes))
        ws_ish.add_image(img_ish, "B5")
    except Exception as e:
        ws_ish.cell(row=5, column=2, value=f"Error al generar diagrama Ishikawa: {e}").font = font_italic

    # ----------------------------------------------------
    # Hojas: 2. Matriz MEFI
    # ----------------------------------------------------
    ws_mefi = cast(Worksheet, wb.create_sheet(title="Matriz MEFI (Interno)"))
    cast(Any, ws_mefi.views.sheetView)[0].showGridLines = True
    create_title_banner(ws_mefi, "MATRIZ DE EVALUACIÓN DE FACTORES INTERNOS (MEFI)", 10)

    ws_mefi.merge_cells("B5:E5")
    f_hdr = cast(Any, ws_mefi.cell(row=5, column=2))
    f_hdr.value = "FORTALEZAS (F) - ANÁLISIS INTERNO"
    f_hdr.font = font_bold_white
    f_hdr.fill = fill_fortaleza_hdr
    f_hdr.alignment = align_center

    ws_mefi.merge_cells("G5:J5")
    d_hdr = cast(Any, ws_mefi.cell(row=5, column=7))
    d_hdr.value = "DEBILIDADES (D) - ANÁLISIS INTERNO"
    d_hdr.font = font_bold_white
    d_hdr.fill = fill_debilidad_hdr
    d_hdr.alignment = align_center

    headers_f = ["Factor Fortaleza", "Peso", "Calificación", "Punt. Ponderada"]
    headers_d = ["Factor Debilidad", "Peso", "Calificación", "Punt. Ponderada"]

    for idx, h in enumerate(headers_f):
        cell = ws_mefi.cell(row=6, column=2 + idx, value=h)
        cell.font = font_bold
        cell.fill = fill_light_gray
        cell.alignment = align_center
        cell.border = border_all_thin

    for idx, h in enumerate(headers_d):
        cell = ws_mefi.cell(row=6, column=7 + idx, value=h)
        cell.font = font_bold
        cell.fill = fill_light_gray
        cell.alignment = align_center
        cell.border = border_all_thin

    list_f = foda_data.get("F", [])
    list_d = foda_data.get("D", [])
    max_rows_mefi = max(len(list_f), len(list_d))

    for r in range(max_rows_mefi):
        row_num = 7 + r
        # Fortaleza
        if r < len(list_f):
            item = list_f[r]
            c1 = ws_mefi.cell(row=row_num, column=2, value=item.get("Factor", ""))
            c2 = ws_mefi.cell(row=row_num, column=3, value=item.get("Peso", 0.0))
            c3 = ws_mefi.cell(row=row_num, column=4, value=item.get("Calificación", 0.0))
            c4 = ws_mefi.cell(row=row_num, column=5, value=f"=C{row_num}*D{row_num}")

            c1.font = font_regular; c1.alignment = align_wrap_left; c1.border = border_all_thin; c1.fill = fill_fortaleza_light
            c2.font = font_regular; c2.alignment = align_right;     c2.border = border_all_thin; c2.number_format = "0.000"
            c3.font = font_regular; c3.alignment = align_right;     c3.border = border_all_thin; c3.number_format = "0.0"
            c4.font = font_regular; c4.alignment = align_right;     c4.border = border_all_thin; c4.number_format = "0.00"
        else:
            for col_c in range(2, 6):
                ws_mefi.cell(row=row_num, column=col_c).border = border_all_thin

        # Debilidad
        if r < len(list_d):
            item = list_d[r]
            c1 = ws_mefi.cell(row=row_num, column=7, value=item.get("Factor", ""))
            c2 = ws_mefi.cell(row=row_num, column=8, value=item.get("Peso", 0.0))
            c3 = ws_mefi.cell(row=row_num, column=9, value=item.get("Calificación", 0.0))
            c4 = ws_mefi.cell(row=row_num, column=10, value=f"=H{row_num}*I{row_num}")

            c1.font = font_regular; c1.alignment = align_wrap_left; c1.border = border_all_thin; c1.fill = fill_debilidad_light
            c2.font = font_regular; c2.alignment = align_right;     c2.border = border_all_thin; c2.number_format = "0.000"
            c3.font = font_regular; c3.alignment = align_right;     c3.border = border_all_thin; c3.number_format = "0.0"
            c4.font = font_regular; c4.alignment = align_right;     c4.border = border_all_thin; c4.number_format = "0.00"
        else:
            for col_c in range(7, 11):
                ws_mefi.cell(row=row_num, column=col_c).border = border_all_thin

    tot_row_mefi = 7 + max_rows_mefi

    # Subtotales
    ws_mefi.cell(row=tot_row_mefi, column=2, value="SUBTOTAL FORTALEZAS").font = font_bold
    ws_mefi.cell(row=tot_row_mefi, column=2).border = border_total_row

    t_f_w = ws_mefi.cell(row=tot_row_mefi, column=3, value=f"=SUM(C7:C{tot_row_mefi-1})")
    t_f_w.font = font_bold; t_f_w.alignment = align_right; t_f_w.border = border_total_row; t_f_w.number_format = "0.000"
    ws_mefi.cell(row=tot_row_mefi, column=4).border = border_total_row

    t_f_s = ws_mefi.cell(row=tot_row_mefi, column=5, value=f"=SUM(E7:E{tot_row_mefi-1})")
    t_f_s.font = font_bold; t_f_s.alignment = align_right; t_f_s.border = border_total_row; t_f_s.number_format = "0.00"

    ws_mefi.cell(row=tot_row_mefi, column=7, value="SUBTOTAL DEBILIDADES").font = font_bold
    ws_mefi.cell(row=tot_row_mefi, column=7).border = border_total_row

    t_d_w = ws_mefi.cell(row=tot_row_mefi, column=8, value=f"=SUM(H7:H{tot_row_mefi-1})")
    t_d_w.font = font_bold; t_d_w.alignment = align_right; t_d_w.border = border_total_row; t_d_w.number_format = "0.000"
    ws_mefi.cell(row=tot_row_mefi, column=9).border = border_total_row

    t_d_s = ws_mefi.cell(row=tot_row_mefi, column=10, value=f"=SUM(J7:J{tot_row_mefi-1})")
    t_d_s.font = font_bold; t_d_s.alignment = align_right; t_d_s.border = border_total_row; t_d_s.number_format = "0.00"

    mefi_final_row = tot_row_mefi + 2
    ws_mefi.merge_cells(start_row=mefi_final_row, start_column=2, end_row=mefi_final_row, end_column=4)
    ws_mefi.cell(row=mefi_final_row, column=2, value="ÍNDICE DE VALORACIÓN INTERNA TOTAL (MEFI)").font = font_bold
    ws_mefi.cell(row=mefi_final_row, column=2).alignment = align_left
    ws_mefi.cell(row=mefi_final_row, column=2).fill = fill_light_gray
    for col_c in range(2, 5):
        ws_mefi.cell(row=mefi_final_row, column=col_c).border = border_all_thin
        ws_mefi.cell(row=mefi_final_row, column=col_c).fill = fill_light_gray

    mefi_val = ws_mefi.cell(row=mefi_final_row, column=5, value=f"=E{tot_row_mefi}+J{tot_row_mefi}")
    mefi_val.font = Font(name=font_family, size=11, bold=True, color="00F2FF")
    mefi_val.fill = fill_navy
    mefi_val.alignment = align_right
    mefi_val.border = border_all_thin
    mefi_val.number_format = "0.00"

    # ----------------------------------------------------
    # Hojas: 3. Matriz MEFE
    # ----------------------------------------------------
    ws_mefe = cast(Worksheet, wb.create_sheet(title="Matriz MEFE (Externo)"))
    cast(Any, ws_mefe.views.sheetView)[0].showGridLines = True
    create_title_banner(ws_mefe, "MATRIZ DE EVALUACIÓN DE FACTORES EXTERNOS (MEFE)", 10)

    ws_mefe.merge_cells("B5:E5")
    o_hdr = cast(Any, ws_mefe.cell(row=5, column=2))
    o_hdr.value = "OPORTUNIDADES (O) - ANÁLISIS EXTERNO"
    o_hdr.font = font_bold_white
    o_hdr.fill = fill_oportunidad_hdr
    o_hdr.alignment = align_center

    ws_mefe.merge_cells("G5:J5")
    a_hdr = cast(Any, ws_mefe.cell(row=5, column=7))
    a_hdr.value = "AMENAZAS (A) - ANÁLISIS EXTERNO"
    a_hdr.font = font_bold_white
    a_hdr.fill = fill_amenaza_hdr
    a_hdr.alignment = align_center

    headers_o = ["Factor Oportunidad", "Peso", "Calificación", "Punt. Ponderada"]
    headers_a = ["Factor Amenaza", "Peso", "Calificación", "Punt. Ponderada"]

    for idx, h in enumerate(headers_o):
        cell = ws_mefe.cell(row=6, column=2 + idx, value=h)
        cell.font = font_bold
        cell.fill = fill_light_gray
        cell.alignment = align_center
        cell.border = border_all_thin

    for idx, h in enumerate(headers_a):
        cell = ws_mefe.cell(row=6, column=7 + idx, value=h)
        cell.font = font_bold
        cell.fill = fill_light_gray
        cell.alignment = align_center
        cell.border = border_all_thin

    list_o = foda_data.get("O", [])
    list_a = foda_data.get("A", [])
    max_rows_mefe = max(len(list_o), len(list_a))

    for r in range(max_rows_mefe):
        row_num = 7 + r
        # Oportunidad
        if r < len(list_o):
            item = list_o[r]
            c1 = ws_mefe.cell(row=row_num, column=2, value=item.get("Factor", ""))
            c2 = ws_mefe.cell(row=row_num, column=3, value=item.get("Peso", 0.0))
            c3 = ws_mefe.cell(row=row_num, column=4, value=item.get("Calificación", 0.0))
            c4 = ws_mefe.cell(row=row_num, column=5, value=f"=C{row_num}*D{row_num}")

            c1.font = font_regular; c1.alignment = align_wrap_left; c1.border = border_all_thin; c1.fill = fill_oportunidad_light
            c2.font = font_regular; c2.alignment = align_right;     c2.border = border_all_thin; c2.number_format = "0.000"
            c3.font = font_regular; c3.alignment = align_right;     c3.border = border_all_thin; c3.number_format = "0.0"
            c4.font = font_regular; c4.alignment = align_right;     c4.border = border_all_thin; c4.number_format = "0.00"
        else:
            for col_c in range(2, 6):
                ws_mefe.cell(row=row_num, column=col_c).border = border_all_thin

        # Amenaza
        if r < len(list_a):
            item = list_a[r]
            c1 = ws_mefe.cell(row=row_num, column=7, value=item.get("Factor", ""))
            c2 = ws_mefe.cell(row=row_num, column=8, value=item.get("Peso", 0.0))
            c3 = ws_mefe.cell(row=row_num, column=9, value=item.get("Calificación", 0.0))
            c4 = ws_mefe.cell(row=row_num, column=10, value=f"=H{row_num}*I{row_num}")

            c1.font = font_regular; c1.alignment = align_wrap_left; c1.border = border_all_thin; c1.fill = fill_amenaza_light
            c2.font = font_regular; c2.alignment = align_right;     c2.border = border_all_thin; c2.number_format = "0.000"
            c3.font = font_regular; c3.alignment = align_right;     c3.border = border_all_thin; c3.number_format = "0.0"
            c4.font = font_regular; c4.alignment = align_right;     c4.border = border_all_thin; c4.number_format = "0.00"
        else:
            for col_c in range(7, 11):
                ws_mefe.cell(row=row_num, column=col_c).border = border_all_thin

    tot_row_mefe = 7 + max_rows_mefe

    # Subtotales
    ws_mefe.cell(row=tot_row_mefe, column=2, value="SUBTOTAL OPORTUNIDADES").font = font_bold
    ws_mefe.cell(row=tot_row_mefe, column=2).border = border_total_row

    t_o_w = ws_mefe.cell(row=tot_row_mefe, column=3, value=f"=SUM(C7:C{tot_row_mefe-1})")
    t_o_w.font = font_bold; t_o_w.alignment = align_right; t_o_w.border = border_total_row; t_o_w.number_format = "0.000"
    ws_mefe.cell(row=tot_row_mefe, column=4).border = border_total_row

    t_o_s = ws_mefe.cell(row=tot_row_mefe, column=5, value=f"=SUM(E7:E{tot_row_mefe-1})")
    t_o_s.font = font_bold; t_o_s.alignment = align_right; t_o_s.border = border_total_row; t_o_s.number_format = "0.00"

    ws_mefe.cell(row=tot_row_mefe, column=7, value="SUBTOTAL AMENAZAS").font = font_bold
    ws_mefe.cell(row=tot_row_mefe, column=7).border = border_total_row

    t_a_w = ws_mefe.cell(row=tot_row_mefe, column=8, value=f"=SUM(H7:H{tot_row_mefe-1})")
    t_a_w.font = font_bold; t_a_w.alignment = align_right; t_a_w.border = border_total_row; t_a_w.number_format = "0.000"
    ws_mefe.cell(row=tot_row_mefe, column=9).border = border_total_row

    t_a_s = ws_mefe.cell(row=tot_row_mefe, column=10, value=f"=SUM(J7:J{tot_row_mefe-1})")
    t_a_s.font = font_bold; t_a_s.alignment = align_right; t_a_s.border = border_total_row; t_a_s.number_format = "0.00"

    mefe_final_row = tot_row_mefe + 2
    ws_mefe.merge_cells(start_row=mefe_final_row, start_column=2, end_row=mefe_final_row, end_column=4)
    ws_mefe.cell(row=mefe_final_row, column=2, value="ÍNDICE DE VALORACIÓN EXTERNA TOTAL (MEFE)").font = font_bold
    ws_mefe.cell(row=mefe_final_row, column=2).alignment = align_left
    ws_mefe.cell(row=mefe_final_row, column=2).fill = fill_light_gray
    for col_c in range(2, 5):
        ws_mefe.cell(row=mefe_final_row, column=col_c).border = border_all_thin
        ws_mefe.cell(row=mefe_final_row, column=col_c).fill = fill_light_gray

    mefe_val = ws_mefe.cell(row=mefe_final_row, column=5, value=f"=E{tot_row_mefe}+J{tot_row_mefe}")
    mefe_val.font = Font(name=font_family, size=11, bold=True, color="00F2FF")
    mefe_val.fill = fill_navy
    mefe_val.alignment = align_right
    mefe_val.border = border_all_thin
    mefe_val.number_format = "0.00"

    # ----------------------------------------------------
    # Hojas: 4. Estrategias CAME
    # ----------------------------------------------------
    ws_came = cast(Worksheet, wb.create_sheet(title="Estrategias CAME"))
    cast(Any, ws_came.views.sheetView)[0].showGridLines = True
    create_title_banner(ws_came, "FORMULACIÓN ESTRATÉGICA DE RECOMENDACIONES CAME", 6)

    headers_came = ["Postura Estratégica", "Acción CAME", "Enfoque / Estrategia Formulada con IA"]
    for idx, h in enumerate(headers_came):
        cell = ws_came.cell(row=5, column=2 + idx, value=h)
        cell.font = font_bold
        cell.fill = fill_light_gray
        cell.alignment = align_center
        cell.border = border_all_thin

    came_data_list = [
        ("A. Estrategia Ofensiva (F + O)", "EXPLOTAR", came_estrategias.get("A. Estrategia Ofensiva (F + O)", "") or "Estrategia no formulada en esta sesión.", fill_fortaleza_hdr, fill_fortaleza_light),
        ("B. Estrategia de Supervivencia (D + A)", "AFRONTAR", came_estrategias.get("B. Estrategia de Supervivencia (D + A)", "") or "Estrategia no formulada en esta sesión.", fill_debilidad_hdr, fill_debilidad_light),
        ("C. Estrategia Defensiva (F + A)", "MANTENER", came_estrategias.get("C. Estrategia Defensiva (F + A)", "") or "Estrategia no formulada en esta sesión.", fill_oportunidad_hdr, fill_oportunidad_light),
        ("D. Estrategia de Reorientación (D + O)", "CORREGIR", came_estrategias.get("D. Estrategia de Reorientación (D + O)", "") or "Estrategia no formulada en esta sesión.", fill_amenaza_hdr, fill_amenaza_light)
    ]

    r_came_idx = 6
    for posture, action, strategy, fill_hdr, fill_lgt in came_data_list:
        c1 = ws_came.cell(row=r_came_idx, column=2, value=posture)
        c2 = ws_came.cell(row=r_came_idx, column=3, value=action)
        c3 = ws_came.cell(row=r_came_idx, column=4, value=strategy)

        c1.font = font_bold;       c1.alignment = align_left;      c1.border = border_all_thin; c1.fill = fill_lgt
        c2.font = font_bold_white; c2.alignment = align_center;    c2.border = border_all_thin; c2.fill = fill_hdr
        c3.font = font_regular;    c3.alignment = align_wrap_left; c3.border = border_all_thin
        ws_came.merge_cells(start_row=r_came_idx, start_column=4, end_row=r_came_idx, end_column=6)

        for col_c in range(4, 7):
            ws_came.cell(row=r_came_idx, column=col_c).border = border_all_thin

        # Ajuste de altura dinámico
        ws_came.row_dimensions[r_came_idx].height = max(55, len(strategy) // 3)
        r_came_idx += 1

    # ----------------------------------------------------
    # Hojas: 5. Plan de Acción y Mando
    # ----------------------------------------------------
    ws_act = cast(Worksheet, wb.create_sheet(title="Plan de Acción y Mando"))
    cast(Any, ws_act.views.sheetView)[0].showGridLines = True
    create_title_banner(ws_act, "MANDO COMPARTIDO Y PLAN DE ACCIÓN TÁCTICO", 7)

    # Consenso Votos
    ws_act.cell(row=5, column=2, value="VOTACIÓN Y CONSENSO DE POSTURA ESTRATÉGICA").font = font_subtitle
    ws_act.cell(row=5, column=2).fill = fill_navy
    ws_act.merge_cells("B5:D5")
    for col_c in range(2, 5):
        ws_act.cell(row=5, column=col_c).border = border_all_thin
        ws_act.cell(row=5, column=col_c).fill = fill_navy

    ws_act.cell(row=6, column=2, value="Línea / Postura").font = font_bold
    ws_act.cell(row=6, column=2).fill = fill_light_gray
    ws_act.cell(row=6, column=2).border = border_all_thin

    ws_act.cell(row=6, column=3, value="Votos Recibidos").font = font_bold
    ws_act.cell(row=6, column=3).fill = fill_light_gray
    ws_act.cell(row=6, column=3).border = border_all_thin

    ws_act.cell(row=6, column=4, value="Porcentaje").font = font_bold
    ws_act.cell(row=6, column=4).fill = fill_light_gray
    ws_act.cell(row=6, column=4).border = border_all_thin

    posturas_votos = [
        ("FO - Estrategia Ofensiva", votos.get("FO", 0)),
        ("DO - Estrategia de Reorientación", votos.get("DO", 0)),
        ("FA - Estrategia Defensiva", votos.get("FA", 0)),
        ("DA - Estrategia de Supervivencia", votos.get("DA", 0))
    ]

    for idx, (posture_name, vote_count) in enumerate(posturas_votos):
        r = 7 + idx
        ws_act.cell(row=r, column=2, value=posture_name).font = font_regular
        ws_act.cell(row=r, column=2).border = border_all_thin

        ws_act.cell(row=r, column=3, value=vote_count).font = font_regular
        ws_act.cell(row=r, column=3).border = border_all_thin
        ws_act.cell(row=r, column=3).alignment = align_right

        pct_cell = ws_act.cell(row=r, column=4, value=f"=IF(SUM(C7:C10)>0, C{r}/SUM(C7:C10), 0)")
        pct_cell.font = font_regular
        pct_cell.border = border_all_thin
        pct_cell.alignment = align_right
        pct_cell.number_format = "0.0%"

    ws_act.cell(row=11, column=2, value="Consenso Alcanzado:").font = font_bold
    ws_act.cell(row=11, column=2).border = border_all_thin
    consenso_val = calcular_consenso(votos, total_votantes)
    ws_act.cell(row=11, column=3, value=consenso_val).font = font_bold
    ws_act.cell(row=11, column=3).border = border_all_thin
    ws_act.merge_cells("C11:D11")
    ws_act.cell(row=11, column=4).border = border_all_thin

    # Plan Acciones
    action_start_row = 13
    ws_act.cell(row=action_start_row, column=2, value="PLAN DE ACCIÓN TÁCTICO").font = font_subtitle
    ws_act.cell(row=action_start_row, column=2).fill = fill_navy
    ws_act.merge_cells(start_row=action_start_row, start_column=2, end_row=action_start_row, end_column=7)
    for col_c in range(2, 8):
        ws_act.cell(row=action_start_row, column=col_c).border = border_all_thin
        ws_act.cell(row=action_start_row, column=col_c).fill = fill_navy

    headers_actions = ["Línea de Acción / Iniciativa", "Prioridad", "Responsable", "Fecha Límite", "Recursos Asignados", "Estatus / Estado"]
    for idx, h in enumerate(headers_actions):
        cell = ws_act.cell(row=action_start_row + 1, column=2 + idx, value=h)
        cell.font = font_bold
        cell.fill = fill_light_gray
        cell.alignment = align_center
        cell.border = border_all_thin

    acciones_list = acciones or []
    act_row_idx = action_start_row + 2
    if not acciones_list:
        ws_act.cell(row=act_row_idx, column=2, value="No se han registrado iniciativas en el Plan de Acción.").font = font_italic
        ws_act.merge_cells(start_row=act_row_idx, start_column=2, end_row=act_row_idx, end_column=7)
        for col_c in range(2, 8):
            ws_act.cell(row=act_row_idx, column=col_c).border = border_all_thin
    else:
        for item in acciones_list:
            c1 = ws_act.cell(row=act_row_idx, column=2, value=item.get("iniciativa", ""))
            c2 = ws_act.cell(row=act_row_idx, column=3, value=item.get("prioridad", ""))
            c3 = ws_act.cell(row=act_row_idx, column=4, value=item.get("responsable", ""))
            c4 = ws_act.cell(row=act_row_idx, column=5, value=str(item.get("fecha_limite", "")))
            c5 = ws_act.cell(row=act_row_idx, column=6, value=item.get("recursos", ""))
            c6 = ws_act.cell(row=act_row_idx, column=7, value=item.get("estado", ""))

            c1.font = font_regular; c1.alignment = align_wrap_left; c1.border = border_all_thin
            c2.font = font_regular; c2.alignment = align_center;    c2.border = border_all_thin
            c3.font = font_regular; c3.alignment = align_left;      c3.border = border_all_thin
            c4.font = font_regular; c4.alignment = align_center;    c4.border = border_all_thin
            c5.font = font_regular; c5.alignment = align_wrap_left; c5.border = border_all_thin
            c6.font = font_bold;    c6.alignment = align_center;    c6.border = border_all_thin

            priority = item.get("prioridad", "").upper()
            if "ALTA" in priority:
                c2.font = Font(name=font_family, size=11, bold=True, color="FF0055")
            elif "MEDIA" in priority:
                c2.font = Font(name=font_family, size=11, bold=True, color="FFAA00")
            else:
                c2.font = Font(name=font_family, size=11, color="2E7D32")

            status = item.get("estado", "").upper()
            if "PENDIENTE" in status or "NO INICIADO" in status:
                c6.font = Font(name=font_family, size=11, bold=True, color="FF0055")
            elif "PROGRESO" in status or "DESARROLLO" in status:
                c6.font = Font(name=font_family, size=11, bold=True, color="FFAA00")
            else:
                c6.font = Font(name=font_family, size=11, bold=True, color="2E7D32")

            act_row_idx += 1

    # ----------------------------------------------------
    # Hojas: 6. Bitácora de Operaciones
    # ----------------------------------------------------
    ws_bit = cast(Worksheet, wb.create_sheet(title="Bitácora de Operaciones"))
    cast(Any, ws_bit.views.sheetView)[0].showGridLines = True
    create_title_banner(ws_bit, "BITÁCORA TÁCTICA DE OPERACIONES (EVENT LOG)", 4)

    ws_bit.cell(row=5, column=2, value="Índice").font = font_bold
    ws_bit.cell(row=5, column=2).fill = fill_light_gray
    ws_bit.cell(row=5, column=2).border = border_all_thin
    ws_bit.cell(row=5, column=2).alignment = align_center

    ws_bit.cell(row=5, column=3, value="Hora de Registro").font = font_bold
    ws_bit.cell(row=5, column=3).fill = fill_light_gray
    ws_bit.cell(row=5, column=3).border = border_all_thin
    ws_bit.cell(row=5, column=3).alignment = align_center

    ws_bit.cell(row=5, column=4, value="Evento / Acción").font = font_bold
    ws_bit.cell(row=5, column=4).fill = fill_light_gray
    ws_bit.cell(row=5, column=4).border = border_all_thin
    ws_bit.cell(row=5, column=4).alignment = align_left

    bitacora_list = bitacora or []
    r_bit_idx = 6
    if not bitacora_list:
        ws_bit.cell(row=r_bit_idx, column=2, value="-").font = font_italic; ws_bit.cell(row=r_bit_idx, column=2).border = border_all_thin; ws_bit.cell(row=r_bit_idx, column=2).alignment = align_center
        ws_bit.cell(row=r_bit_idx, column=3, value="-").font = font_italic; ws_bit.cell(row=r_bit_idx, column=3).border = border_all_thin; ws_bit.cell(row=r_bit_idx, column=3).alignment = align_center
        ws_bit.cell(row=r_bit_idx, column=4, value="No hay eventos registrados en la bitácora.").font = font_italic; ws_bit.cell(row=r_bit_idx, column=4).border = border_all_thin
    else:
        for idx, item in enumerate(bitacora_list):
            c1 = ws_bit.cell(row=r_bit_idx, column=2, value=idx + 1)
            c2 = ws_bit.cell(row=r_bit_idx, column=3, value=item.get("hora", ""))
            c3 = ws_bit.cell(row=r_bit_idx, column=4, value=item.get("evento", ""))

            c1.font = font_regular; c1.alignment = align_center; c1.border = border_all_thin
            c2.font = font_regular; c2.alignment = align_center; c2.border = border_all_thin
            c3.font = font_regular; c3.alignment = align_left;   c3.border = border_all_thin
            r_bit_idx += 1

    # ----------------------------------------------------
    # Anchos de Columna Dinámicos
    # ----------------------------------------------------
    for sheet_raw in wb.worksheets:
        sheet = cast(Worksheet, sheet_raw)
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    if cell.row in [2, 3]:
                        continue
                    cell_lines = str(cell.value).split('\n')
                    for line in cell_lines:
                        if line.startswith('='):
                            max_len = max(max_len, 10)
                        else:
                            max_len = max(max_len, len(line))
            col_letter = get_column_letter(cast(int, col[0].column))

            if sheet.title == "Resumen Ejecutivo" and col_letter == 'B':
                sheet.column_dimensions[col_letter].width = 30
            elif sheet.title == "Resumen Ejecutivo" and col_letter == 'C':
                sheet.column_dimensions[col_letter].width = 65
            elif sheet.title == "Estrategias CAME" and col_letter == 'D':
                sheet.column_dimensions[col_letter].width = 75
            else:
                sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 11), 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@st.fragment
def renderizar_panel_exportacion_informe(foda_data, x, y, prob_exito, perfil, tipo_ana, f_inicio, f_fin):
    fmt = st.radio("Formato de Exportación", ["PDF", "EXCEL"], horizontal=True, key="fmt_exp_informe_frag")
    if fmt == "PDF":
        if not st.session_state.get("pdf_ready_informe"):
            if st.button("🔄 PREPARAR DESCARGA DE REPORTE PDF", use_container_width=True, key="btn_prep_pdf_informe_frag"):
                progress_bar = st.progress(0.0, text="Inicializando motor de compilación PDF...")
                try:
                    pdf_bytes = crear_pdf_final(
                        foda_dict=foda_data,
                        texto_ia=st.session_state.ultimo_resultado,
                        x=x,
                        y=y,
                        perfil=perfil,
                        tipo_ana=tipo_ana,
                        fecha_i=f_inicio,
                        fecha_t=f_fin,
                        prob_exito=prob_exito,
                        bitacora=st.session_state.bitacora,
                        acciones=st.session_state.acciones,
                        votos=st.session_state.votos,
                        total_votantes=st.session_state.total_votantes,
                        came_estrategias=st.session_state.came_estrategias,
                        stress_int=st.session_state.stress_int,
                        stress_ext=st.session_state.stress_ext,
                        progress_cb=make_progress_cb(progress_bar)
                    )
                    progress_bar.progress(1.0, text="Reporte PDF Compilado con Éxito.")
                    time.sleep(0.4)
                    progress_bar.empty()
                    st.session_state.pdf_bytes_informe = pdf_bytes
                    st.session_state.pdf_ready_informe = True
                    st.rerun() # Usar rerun() dentro del fragmento para actualizar la UI del fragmento de forma reactiva
                except Exception as e:
                    st.error(f"Error al generar el PDF: {str(e)}")
        else:
            if "pdf_bytes_informe" in st.session_state:
                st.download_button(
                    label="📥 DESCARGAR REPORTE PDF",
                    data=st.session_state.pdf_bytes_informe,
                    file_name="analisis_foda_s5.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="btn_descargar_pdf_informe_real"
                )
            else:
                st.session_state.pdf_ready_informe = False
                st.rerun()
    else:
        try:
            # Exportar Excel directamente con formato premium táctico
            excel_data = crear_excel_estilizado(
                foda_data=foda_data,
                x=x,
                y=y,
                prob_exito=prob_exito,
                perfil=perfil,
                tipo_ana=tipo_ana,
                f_inicio=f_inicio,
                f_fin=f_fin,
                bitacora=st.session_state.bitacora,
                votos=st.session_state.votos,
                total_votantes=st.session_state.total_votantes,
                came_estrategias=st.session_state.came_estrategias,
                alert_level=st.session_state.alert_level,
                ultimo_resultado=st.session_state.ultimo_resultado,
                acciones=st.session_state.acciones
            )
            st.download_button(
                "📥 DESCARGAR EXCEL",
                data=excel_data,
                file_name="analisis_foda_s5.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_descargar_excel_informe_frag",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"Error al generar Excel: {str(e)}")

@st.fragment
def renderizar_panel_exportacion_mando(foda_data_mando, x, y, prob_exito, p_val, t_val, fi_val, ff_val):
    if not st.session_state.get("pdf_ready_mando"):
        if st.button("🔄 PREPARAR DESCARGA DE REPORTE PDF", use_container_width=True, key="btn_prep_pdf_mando_frag"):
            progress_bar = st.progress(0.0, text="Inicializando motor de compilación PDF...")
            try:
                pdf_bytes_mando = crear_pdf_final(
                    foda_dict=foda_data_mando,
                    texto_ia=st.session_state.ultimo_resultado,
                    x=x,
                    y=y,
                    perfil=p_val,
                    tipo_ana=t_val,
                    fecha_i=fi_val,
                    fecha_t=ff_val,
                    prob_exito=prob_exito,
                    bitacora=st.session_state.bitacora,
                    acciones=st.session_state.acciones,
                    votos=st.session_state.votos,
                    total_votantes=st.session_state.total_votantes,
                    came_estrategias=st.session_state.came_estrategias,
                    stress_int=st.session_state.stress_int,
                    stress_ext=st.session_state.stress_ext,
                    progress_cb=make_progress_cb(progress_bar)
                )
                progress_bar.progress(1.0, text="Reporte PDF Compilado con Éxito.")
                time.sleep(0.4)
                progress_bar.empty()
                st.session_state.pdf_bytes_mando = pdf_bytes_mando
                st.session_state.pdf_ready_mando = True
                st.rerun()
            except Exception as e:
                st.error(f"Error al generar el PDF: {str(e)}")
    else:
        if "pdf_bytes_mando" in st.session_state:
            st.download_button(
                label="📥 DESCARGAR REPORTE ESTRATÉGICO PDF",
                data=st.session_state.pdf_bytes_mando,
                file_name="analisis_foda_s5.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="btn_descargar_pdf_mando_real"
            )
        else:
            st.session_state.pdf_ready_mando = False
            st.rerun()

init_db()

if 'data_foda' not in st.session_state:
    st.session_state.data_foda = {
        "F": [{"Factor": "Despliegue inmediato y movilidad", "Peso": 0.25, "Calificación": 4.0}],
        "D": [{"Factor": "Sobrecarga por tareas ajenas", "Peso": 0.25, "Calificación": 2.0}],
        "O": [{"Factor": "Mejora de medios operativos", "Peso": 0.25, "Calificación": 3.0}],
        "A": [{"Factor": "Ciberataques contra la institución", "Peso": 0.25, "Calificación": 1.0}]
    }
if 'ultimo_resultado' not in st.session_state:
    st.session_state.ultimo_resultado = ""
if 'pdf_ready_informe' not in st.session_state:
    st.session_state.pdf_ready_informe = False
if 'pdf_ready_mando' not in st.session_state:
    st.session_state.pdf_ready_mando = False
if "last_valid_tab" not in st.session_state:
    st.session_state["last_valid_tab"] = "MATRICES PONDERADAS"
if "rol" not in st.session_state:
    st.session_state.rol = "Analista"
if "bitacora" not in st.session_state:
    st.session_state.bitacora = []
if "votos" not in st.session_state:
    st.session_state.votos = {"FO": 0, "DO": 0, "FA": 0, "DA": 0}
if "total_votantes" not in st.session_state:
    st.session_state.total_votantes = 10
if "acciones" not in st.session_state:
    st.session_state.acciones = []
if "alert_level" not in st.session_state:
    st.session_state.alert_level = "AMARILLO"
if "came_estrategias" not in st.session_state:
    st.session_state.came_estrategias = {
        "A. Estrategia Ofensiva (F + O)": "",
        "B. Estrategia de Supervivencia (D + A)": "",
        "C. Estrategia Defensiva (F + A)": "",
        "D. Estrategia de Reorientación (D + O)": ""
    }

# Modificadores de estrés en sesión (para conservar el estado entre pestañas si es necesario)
if "stress_int" not in st.session_state:
    st.session_state.stress_int = 0.0
if "stress_ext" not in st.session_state:
    st.session_state.stress_ext = 0.0

if "perfil_val" not in st.session_state:
    st.session_state.perfil_val = "Sin perfil"
if "tipo_ana_val" not in st.session_state:
    st.session_state.tipo_ana_val = "Análisis General"
from datetime import date, timedelta
if "f_inicio_val" not in st.session_state:
    st.session_state.f_inicio_val = date.today()
if "f_fin_val" not in st.session_state:
    st.session_state.f_fin_val = date.today() + timedelta(days=60)

if "prompt_foda" not in st.session_state:
    st.session_state.prompt_foda = """ESTÁS ACTUANDO EN EL MODO: SISTEMA DE INTELIGENCIA DE ESTADO MAYOR FODA S5.
PERFIL DE OPERACIONES: {perfil}
CATEGORÍA DEL ANÁLISIS: {tipo_ana}
VENTANA TEMPORAL DE EVALUACIÓN: {fecha_i} a {fecha_t}
COORDENADAS ESTRATÉGICAS ACTUALES: Vector X={x}, Y={y}

DIRECTIVAS DE FORMATO Y CONTENIDO (ESTRICTAS Y MANDATORIAS):
1. Redacta de forma altamente profesional, militar y geopolítica.
2. El reporte debe estructurarse EXACTAMENTE en los siguientes cuatro rubros principales (escribe cada rubro en mayúsculas, en una línea separada, sin corchetes):
   SITUACIÓN GENERAL
   AMENAZAS
   RIESGOS
   LÍNEAS DE ACCIÓN

3. Para el rubro "SITUACIÓN GENERAL", redacta exclusivamente un único párrafo descriptivo en prosa, sin numeración ni viñetas.
4. Para los rubros "AMENAZAS", "RIESGOS", y "LÍNEAS DE ACCIÓN", presenta los resultados en forma de lista numerada con números arábigos (1., 2., 3., etc.).
5. Asegúrate de que el texto fluya de manera uniforme para lucir impecable al mostrarse justificado (text-align: justify).
6. Utiliza emojis (como 🔵 para SITUACIÓN GENERAL; 🟡 para AMENAZAS; 🔴 para RIESGOS; 🟢 para LÍNEAS DE ACCIÓN) ÚNICAMENTE como prefijo inicial de sus respectivos encabezados. Queda estrictamente prohibido utilizar emojis, iconos o símbolos especiales dentro del cuerpo de los párrafos o en las oraciones de las listas.
7. Mantén un interlineado sencillo (un salto de línea normal) dentro de los párrafos y listas. No agregues líneas vacías adicionales dentro de una sección.
8. Agrega exactamente un doble salto de línea (una línea vacía) antes de cada nuevo rubro o encabezado con Emoji para separar las secciones.
9. DIRECTIVA DE IDIOMA CRÍTICA: Redacta la respuesta única y exclusivamente en español castellano profesional y militar. Queda terminantemente prohibido utilizar palabras en inglés, siglas en inglés, anglicismos o traducciones parciales (por ejemplo, usa 'nuestro' o 'nuestros' en lugar de 'our'). Todo el texto generado debe ser 100% en español puro."""

if "prompt_came" not in st.session_state:
    st.session_state.prompt_came = """ESTÁS ACTUANDO EN EL MODO: SISTEMA DE PLANIFICACIÓN ESTRATÉGICA DE ESTADO MAYOR FODA S5.
LÍNEA ESTRATÉGICA CAME A FORMULAR: {estrategia_tipo}
OBJETIVO PRINCIPAL: {objetivo}
LÓGICA METODOLÓGICA: {logica}

DIRECTIVAS DE CONTENIDO (ESTRICTAS Y MANDATORIAS):
1. Redacta de forma altamente profesional, táctica y militar.
2. Utiliza un tono ejecutivo, enfocado a la toma de decisiones estratégicas.
3. Analiza los factores FODA activos proporcionados.
4. Estructura el reporte exactamente de la siguiente manera:
   - Punto 1: Escribe "1. [Primer Factor (e.g. Fortalezas o Debilidades)]:" en una línea, y en la siguiente línea redacta un análisis en prosa justificando este factor.
   - Punto 2: Escribe "2. [Segundo Factor (e.g. Oportunidades o Amenazas)]:" en una línea, y en la siguiente línea redacta un análisis en prosa justificando este factor.
   - Punto 3: Escribe "3. LÍNEAS DE ACCIÓN:" en una línea, y a continuación propón exactamente 4 líneas de acción operacionales y accionables, numeradas secuencialmente iniciando desde 1 como "1. [Línea 1]", "2. [Línea 2]", "3. [Línea 3]" y "4. [Línea 4]".
5. No uses viñetas como guiones o asteriscos en la sección 3.
6. Asegúrate de que el texto fluya de manera uniforme para lucir impecable al mostrarse justificado (text-align: justify).
7. No incluyas explicaciones preliminares, saludos, ni cierres.
8. NO utilices ningún tipo de ícono, símbolo especial o emoji. El texto debe estar limpio de emojis y contener únicamente texto plano y números arábigos.
9. Mantén un interlineado sencillo (sin líneas vacías redundantes dentro de una misma sección).
10. DIRECTIVA DE IDIOMA CRÍTICA: Redacta la respuesta única y exclusivamente en español castellano profesional y militar. Queda terminantemente prohibido utilizar palabras en inglés, siglas en inglés, anglicismos o traducciones parciales (por ejemplo, usa 'nuestro' o 'nuestros' en lugar de 'our'). Todo el texto generado debe ser 100% en español puro."""

def registrar_evento(evento):
    st.session_state.bitacora.append({
        "hora": datetime.now().strftime("%H:%M:%S"),
        "evento": evento
    })

def calcular_consenso(votos, total_votantes):
    if total_votantes <= 0:
        return "RESULTADO: CONFIGURACIÓN DE VOTANTES INVÁLIDA"
    ganador = max(votos, key=votos.get)
    votos_ganador = votos[ganador]
    if votos_ganador == 0:
        return "RESULTADO: SIN VOTACIÓN ACTIVA"
    porcentaje = (votos_ganador / total_votantes) * 100
    if porcentaje > 60:
        return f"CONSENSO DIRECTIVA: {ganador} ({porcentaje:.1f}%)"
    else:
        return f"SIN CONSENSO: {ganador} CON {porcentaje:.1f}% (REQUERIDO > 60%)"

# =============================================================================
# 5. FUNCIÓN COMPONENTE DE EDICIÓN DE TABLAS (UI)
# =============================================================================
def mostrar_tabla(titulo, tipo):
    st.markdown(f"**{titulo}**")
    
    edited_df = st.data_editor(
        st.session_state[f"df_{tipo}"],
        column_config={
            "Factor": st.column_config.TextColumn(
                "Factor / Criterio Operacional",
                help="Descripción detallada del factor de superioridad o mitigación",
            ),
            "Peso": st.column_config.NumberColumn(
                "Peso Ponderado",
                help="Importancia relativa (0.00 a 1.00). La suma de la matriz interna/externa debe ser 1.00",
                min_value=0.00,
                max_value=1.00,
                step=0.01,
                format="%.3f",
            ),
            "Calificación": st.column_config.NumberColumn(
                "Calificación",
                help="Grado de respuesta o impacto (1.0 a 5.0)",
                min_value=1.0,
                max_value=5.0,
                step=0.1,
                format="%.1f",
            )
        },
        num_rows="dynamic",
        key=f"editor_{tipo}",
        width="stretch",
        height=230
    )
    
    return edited_df

def render_tab_matrices():
    # Cargar DataFrames de sesión
    df_f = st.session_state["df_F"]
    df_d = st.session_state["df_D"]
    df_o = st.session_state["df_O"]
    df_a = st.session_state["df_A"]
    
    # Mensaje de confirmación visual para el operador (usando st.toast para evitar saltos/desplazamientos de página)
    if st.session_state.get("matrices_saved_success"):
        msg_type = st.session_state["matrices_saved_success"]
        if msg_type == "GUARDAR":
            st.toast("🟢 Matrices guardadas y balance general recalculado con éxito.", icon="✅")
        elif msg_type == "MEFI":
            st.toast("🟢 Autonormalización interna (MEFI) realizada y guardada con éxito.", icon="✅")
        elif msg_type == "MEFE":
            st.toast("🟢 Autonormalización externa (MEFE) realizada y guardada con éxito.", icon="✅")
        elif msg_type == "MEFI_RESET":
            st.toast("🔄 Evaluación MEFI reiniciada con pesos balanceados.", icon="🔄")
        elif msg_type == "MEFE_RESET":
            st.toast("🔄 Evaluación MEFE reiniciada con pesos balanceados.", icon="🔄")
        st.session_state["matrices_saved_success"] = None
        
    with st.container(border=True):
        st.markdown('<div class="hud-panel-title">FODA // FACTORES DE INFLUENCIA ESTRATÉGICA PONDERADA</div>', unsafe_allow_html=True)
        
        # Carga de archivos con mapeo NLP inteligente
        archivo_cargado = st.file_uploader("Cargar matriz de factores desde CSV o Excel (Mapeo Inteligente NLP)", type=["csv", "xlsx"])
        if archivo_cargado is not None:
            try:
                if archivo_cargado.name.endswith(".csv"):
                    df_upload = pd.read_csv(archivo_cargado)
                else:
                    df_upload = pd.read_excel(archivo_cargado)
                    
                # Mapear columnas inteligentemente usando nlp_mapper
                mapping = map_columns_to_foda(df_upload)
                
                cargado = False
                for key, col_name in mapping.items():
                    if col_name is not None and col_name in df_upload.columns:
                        sess_key = key[0].upper() # 'fortalezas' -> 'F'
                        
                        valores = []
                        for idx, row in df_upload.iterrows():
                            raw_val = cast(Any, row[col_name])
                            val_factor = str(raw_val).strip()
                            if pd.isna(raw_val) or not val_factor:
                                continue
                            
                            val_peso = 0.25
                            val_calif = 3.0
                            
                            # Buscar si hay columnas específicas de peso/calificación
                            for col in df_upload.columns:
                                norm_col = col.lower()
                                if "peso" in norm_col or "weight" in norm_col:
                                    try:
                                        raw_peso = cast(Any, row[col])
                                        val_peso = float(raw_peso) if not pd.isna(raw_peso) else 0.25
                                    except ValueError:
                                        val_peso = 0.25
                                elif "calif" in norm_col or "rating" in norm_col or "nota" in norm_col:
                                    try:
                                        raw_calif = cast(Any, row[col])
                                        val_calif = float(raw_calif) if not pd.isna(raw_calif) else 3.0
                                    except ValueError:
                                        val_calif = 3.0
                                    
                            valores.append({
                                "Factor": val_factor,
                                "Peso": val_peso,
                                "Calificación": val_calif
                            })
                        
                        if valores:
                            st.session_state[f"df_{sess_key}"] = sanitizar_dataframe(pd.DataFrame(valores))
                            st.session_state.data_foda[sess_key] = st.session_state[f"df_{sess_key}"].to_dict('records')
                            cargado = True
                
                if cargado:
                    st.toast("Mapeo NLP Completado. Factores cargados exitosamente.")
                    registrar_evento("Cargados factores desde archivo con mapeo NLP")
                    st.rerun()
                else:
                    st.error("No se pudieron mapear las columnas del archivo al formato FODA. Asegúrese de que contenga columnas con nombres similares a Fortalezas, Debilidades, Oportunidades o Amenazas.")
            except Exception as e:
                st.error(f"Error al procesar el archivo: {e}")

        # Validación MEFI/MEFE
        sum_mefi = round(df_f['Peso'].sum() + df_d['Peso'].sum(), 3)
        sum_mefe = round(df_o['Peso'].sum() + df_a['Peso'].sum(), 3)
        
        # Iniciar formulario arriba para englobar la validación y los botones
        with st.form("foda_matrices_form"):
            col_v1, col_v2 = st.columns(2)
            norm_mefi_clicked = False
            norm_mefe_clicked = False
            reset_mefi_clicked = False
            reset_mefe_clicked = False
            
            with col_v1:
                if abs(sum_mefi - 1.0) < 0.01:
                    st.success(f"AUDITORÍA MEFI: OK | Ponderación Interna Balanceada ({sum_mefi:.2f} / 1.00)")
                    reset_mefi_clicked = st.form_submit_button("REINICIAR EVALUACIÓN MEFI", width="stretch")
                else:
                    st.error(f"AUDITORÍA MEFI: NO ACTIVADA | La suma de pesos de la matriz interna debe ser 1.00. (Actual: {sum_mefi:.2f})")
                    norm_mefi_clicked = st.form_submit_button("AUTONORMALIZAR MATRIZ INTERNA (MEFI)", width="stretch")
                    
            with col_v2:
                if abs(sum_mefe - 1.0) < 0.01:
                    st.success(f"AUDITORÍA MEFE: OK | Ponderación Externa Balanceada ({sum_mefe:.2f} / 1.00)")
                    reset_mefe_clicked = st.form_submit_button("REINICIAR EVALUACIÓN MEFE", width="stretch")
                else:
                    st.error(f"AUDITORÍA MEFE: NO ACTIVADA | La suma de pesos de la matriz externa debe ser 1.00. (Actual: {sum_mefe:.2f})")
                    norm_mefe_clicked = st.form_submit_button("AUTONORMALIZAR MATRIZ EXTERNA (MEFE)", width="stretch")

            st.markdown("---")
            
            col_m1, col_m2 = st.columns(2)
            with col_m1:
                df_f_new = mostrar_tabla("FORTALEZAS F (Factores Internos de Superioridad)", "F")
                st.markdown("")
                df_d_new = mostrar_tabla("DEBILIDADES D (Vulnerabilidades Internas de Mitigación)", "D")
            with col_m2:
                df_o_new = mostrar_tabla("OPORTUNIDADES O (Escenarios Externos de Explotación)", "O")
                st.markdown("")
                df_a_new = mostrar_tabla("AMENAZAS A (Factores Externos de Contención)", "A")
                
            submitted = st.form_submit_button("GUARDAR Y APLICAR CAMBIOS EN MATRICES", width="stretch")
            
            if submitted:
                st.session_state["df_F"] = sanitizar_dataframe(df_f_new)
                st.session_state["df_D"] = sanitizar_dataframe(df_d_new)
                st.session_state["df_O"] = sanitizar_dataframe(df_o_new)
                st.session_state["df_A"] = sanitizar_dataframe(df_a_new)
                
                st.session_state.data_foda["F"] = st.session_state["df_F"].to_dict('records')
                st.session_state.data_foda["D"] = st.session_state["df_D"].to_dict('records')
                st.session_state.data_foda["O"] = st.session_state["df_O"].to_dict('records')
                st.session_state.data_foda["A"] = st.session_state["df_A"].to_dict('records')
                
                registrar_evento("Guardados y aplicados cambios en las matrices FODA")
                st.session_state["matrices_saved_success"] = "GUARDAR"
                limpiar_analisis_memoria()
                st.rerun()
                
            elif norm_mefi_clicked:
                # Sanitizar primero para no perder cambios del usuario
                df_f_clean = sanitizar_dataframe(df_f_new)
                df_d_clean = sanitizar_dataframe(df_d_new)
                df_o_clean = sanitizar_dataframe(df_o_new)
                df_a_clean = sanitizar_dataframe(df_a_new)
                
                total = df_f_clean['Peso'].sum() + df_d_clean['Peso'].sum()
                if total > 0:
                    df_f_clean['Peso'] = (df_f_clean['Peso'] / total).round(3)
                    df_d_clean['Peso'] = (df_d_clean['Peso'] / total).round(3)
                
                st.session_state["df_F"] = df_f_clean
                st.session_state["df_D"] = df_d_clean
                st.session_state["df_O"] = df_o_clean
                st.session_state["df_A"] = df_a_clean
                
                st.session_state.data_foda["F"] = df_f_clean.to_dict('records')
                st.session_state.data_foda["D"] = df_d_clean.to_dict('records')
                st.session_state.data_foda["O"] = df_o_clean.to_dict('records')
                st.session_state.data_foda["A"] = df_a_clean.to_dict('records')
                
                registrar_evento("Normalizados pesos de matriz interna MEFI conservando cambios")
                st.session_state["matrices_saved_success"] = "MEFI"
                limpiar_analisis_memoria()
                st.rerun()
                
            elif norm_mefe_clicked:
                # Sanitizar primero para no perder cambios del usuario
                df_f_clean = sanitizar_dataframe(df_f_new)
                df_d_clean = sanitizar_dataframe(df_d_new)
                df_o_clean = sanitizar_dataframe(df_o_new)
                df_a_clean = sanitizar_dataframe(df_a_new)
                
                total = df_o_clean['Peso'].sum() + df_a_clean['Peso'].sum()
                if total > 0:
                    df_o_clean['Peso'] = (df_o_clean['Peso'] / total).round(3)
                    df_a_clean['Peso'] = (df_a_clean['Peso'] / total).round(3)
                
                st.session_state["df_F"] = df_f_clean
                st.session_state["df_D"] = df_d_clean
                st.session_state["df_O"] = df_o_clean
                st.session_state["df_A"] = df_a_clean
                
                st.session_state.data_foda["F"] = df_f_clean.to_dict('records')
                st.session_state.data_foda["D"] = df_d_clean.to_dict('records')
                st.session_state.data_foda["O"] = df_o_clean.to_dict('records')
                st.session_state.data_foda["A"] = df_a_clean.to_dict('records')
                
                registrar_evento("Normalizados pesos de matriz externa MEFE conservando cambios")
                st.session_state["matrices_saved_success"] = "MEFE"
                limpiar_analisis_memoria()
                st.rerun()

            elif reset_mefi_clicked:
                df_f_clean = sanitizar_dataframe(df_f_new)
                df_d_clean = sanitizar_dataframe(df_d_new)
                total_count = len(df_f_clean) + len(df_d_clean)
                if total_count > 0:
                    equal_weight = round(1.0 / total_count, 3)
                    df_f_clean['Peso'] = equal_weight
                    df_d_clean['Peso'] = equal_weight
                    diff = round(1.0 - (np.sum(df_f_clean['Peso']) + np.sum(df_d_clean['Peso'])), 3)
                    if len(df_f_clean) > 0:
                        df_f_clean.loc[df_f_clean.index[0], 'Peso'] = round(df_f_clean.loc[df_f_clean.index[0], 'Peso'] + diff, 3)
                    elif len(df_d_clean) > 0:
                        df_d_clean.loc[df_d_clean.index[0], 'Peso'] = round(df_d_clean.loc[df_d_clean.index[0], 'Peso'] + diff, 3)
                
                st.session_state["df_F"] = df_f_clean
                st.session_state["df_D"] = df_d_clean
                st.session_state.data_foda["F"] = df_f_clean.to_dict('records')
                st.session_state.data_foda["D"] = df_d_clean.to_dict('records')
                registrar_evento("Reiniciada evaluación MEFI con pesos balanceados")
                st.session_state["matrices_saved_success"] = "MEFI_RESET"
                limpiar_analisis_memoria()
                st.rerun()

            elif reset_mefe_clicked:
                df_o_clean = sanitizar_dataframe(df_o_new)
                df_a_clean = sanitizar_dataframe(df_a_new)
                total_count = len(df_o_clean) + len(df_a_clean)
                if total_count > 0:
                    equal_weight = round(1.0 / total_count, 3)
                    df_o_clean['Peso'] = equal_weight
                    df_a_clean['Peso'] = equal_weight
                    diff = round(1.0 - (np.sum(df_o_clean['Peso']) + np.sum(df_a_clean['Peso'])), 3)
                    if len(df_o_clean) > 0:
                        df_o_clean.loc[df_o_clean.index[0], 'Peso'] = round(df_o_clean.loc[df_o_clean.index[0], 'Peso'] + diff, 3)
                    elif len(df_a_clean) > 0:
                        df_a_clean.loc[df_a_clean.index[0], 'Peso'] = round(df_a_clean.loc[df_a_clean.index[0], 'Peso'] + diff, 3)
                
                st.session_state["df_O"] = df_o_clean
                st.session_state["df_A"] = df_a_clean
                st.session_state.data_foda["O"] = df_o_clean.to_dict('records')
                st.session_state.data_foda["A"] = df_a_clean.to_dict('records')
                registrar_evento("Reiniciada evaluación MEFE con pesos balanceados")
                st.session_state["matrices_saved_success"] = "MEFE_RESET"
                limpiar_analisis_memoria()
                st.rerun()

# =============================================================================
# 6. INTERFAZ PRINCIPAL DEL COMANDO TÁCTICO FODA S5 (DISEÑO INTEGRAL)
# =============================================================================

# --- INICIALIZACIÓN DE DATOS Y MÉTRICAS GLOBALES ---
for t in ["F", "D", "O", "A"]:
    if f"df_{t}" not in st.session_state:
        st.session_state[f"df_{t}"] = sanitizar_dataframe(pd.DataFrame(st.session_state.data_foda[t]))

df_f_raw = st.session_state["df_F"]
df_d_raw = st.session_state["df_D"]
df_o_raw = st.session_state["df_O"]
df_a_raw = st.session_state["df_A"]

df_f = sanitizar_dataframe(df_f_raw)
df_d = sanitizar_dataframe(df_d_raw)
df_o = sanitizar_dataframe(df_o_raw)
df_a = sanitizar_dataframe(df_a_raw)

x, y = calcular_scores(df_f, df_d, df_o, df_a)
scores = calcular_mefi_mefe(df_f, df_d, df_o, df_a)
estado_actual = clasificacion(scores)

# Simulación Monte Carlo rápida para calcular probabilidad de éxito en tiempo real
resultados_mc = simulacion_montecarlo(df_f, df_d, df_o, df_a, st.session_state.stress_int, st.session_state.stress_ext)
prob_exito = probabilidad_exito(resultados_mc)


# Nivel de alerta dinámico basado en éxito, amenazas críticas y modificadores de estrés
hay_amenaza_critica = not df_a.empty and (df_a['Calificación'] >= 4.5).any()
stress_total = abs(st.session_state.stress_int) + abs(st.session_state.stress_ext)

if prob_exito < 40.0 or hay_amenaza_critica or stress_total > 20.0:
    st.session_state.alert_level = "ROJO"
elif prob_exito < 70.0 or stress_total > 5.0:
    st.session_state.alert_level = "AMARILLO"
else:
    st.session_state.alert_level = "VERDE"

# --- OVERLAY ---
st.markdown('<div class="hud-overlay"></div>', unsafe_allow_html=True)

# --- CABECERA HUD ---
ahora = datetime.now().strftime("%H:%M:%S")
clase_alerta = ""
if st.session_state.alert_level == "ROJO":
    clase_alerta = "critical"
elif st.session_state.alert_level == "VERDE":
    clase_alerta = "normal"

st.markdown(f"""
<div class="hud-header">
    <div class="left-box" style="display: flex; align-items: center; gap: 12px;">
        <button id="tactical-menu-btn" 
                onclick="const btnStreamlit = document.querySelector('button[data-testid=stSidebarCollapseButton]'); if (btnStreamlit) btnStreamlit.click(); this.style.display = 'none';"
                style="display: none; background: rgba(2, 6, 23, 0.85); border: 1px solid #00f2fe; color: #00f2fe; padding: 4px 10px; border-radius: 4px; font-family: 'Orbitron', sans-serif; font-size: 11px; cursor: pointer; text-shadow: 0 0 5px #00f2fe; box-shadow: 0 0 5px rgba(0, 242, 254, 0.3); transition: all 0.2s ease; letter-spacing: 1px; font-weight: 700;">
            ▶ EXPANDIR
        </button>
    </div>
    <div class="center-box">HERRAMIENTA DE ANÁLISIS EXCLUSIVA PARA JOPNAV/S-5</div>
    <div class="right-box">
        <div class="alert-level {clase_alerta}">ALERTA: {st.session_state.alert_level}</div>
        <div id="hud-clock">{ahora}</div>
    </div>
</div>
""", unsafe_allow_html=True)

# Inyección de script dinámico en el parent window
st.iframe("""
<script>
    (function() {
        const parentDoc = window.parent.document;
        // Remover versiones anteriores del script
        const oldScripts = parentDoc.querySelectorAll("script[id^='hud-clock-script']");
        oldScripts.forEach(s => s.remove());

        const script = parentDoc.createElement("script");
        script.id = "hud-clock-script-v5";
        script.textContent = `
            (function() {
                if (window.__hud_clock_timer) clearInterval(window.__hud_clock_timer);

                window.__hud_tick = function() {
                    const el = document.getElementById("hud-clock");
                    if (el) {
                        const now = new Date();
                        const h = String(now.getHours()).padStart(2, '0');
                        const m = String(now.getMinutes()).padStart(2, '0');
                        const s = String(now.getSeconds()).padStart(2, '0');
                        el.textContent = h + ":" + m + ":" + s;
                    }
                    
                    const sidebar = document.querySelector('section[data-testid="stSidebar"]');
                    if (sidebar) {
                        const buttons = sidebar.querySelectorAll('button');
                        buttons.forEach(b => {
                            if (b.textContent && b.textContent.includes("BORRAR HISTORIAL")) {
                                b.classList.add("btn-destructive");
                            }
                        });
                    }

                    const btn = document.getElementById("tactical-menu-btn");
                    if (btn) {
                        const isCollapsed = !sidebar || sidebar.getAttribute("data-collapsed") === "true";
                        if (isCollapsed) {
                            btn.innerHTML = "▶ EXPANDIR";
                            btn.style.borderColor = "#00f2fe";
                            btn.style.color = "#00f2fe";
                            btn.style.textShadow = "0 0 5px #00f2fe";
                            btn.style.boxShadow = "0 0 5px rgba(0, 242, 254, 0.3)";
                            btn.style.display = "inline-block";
                        } else {
                            btn.style.display = "none";
                        }
                    }
                };
                window.__hud_clock_timer = setInterval(window.__hud_tick, 1000);
                window.__hud_tick();

                // --- PREVENCIÓN DE SCROLL EN EDICIÓN (SIN DESALINEACIÓN) ---
                // Bloquea gestos de scroll únicamente cuando se edita para evitar que la celda flote
                const preventDefault = (e) => {
                    const active = document.activeElement;
                    if (active && (active.tagName === 'INPUT' || active.tagName === 'TEXTAREA')) {
                        if (e.target === active) return;
                        e.preventDefault();
                    }
                };

                const preventKeys = (e) => {
                    const active = document.activeElement;
                    if (active && (active.tagName === 'INPUT' || active.tagName === 'TEXTAREA')) {
                        const keys = ['PageUp', 'PageDown'];
                        if (keys.includes(e.code)) {
                            e.preventDefault();
                        }
                    }
                };

                if (window.__hud_wheel_handler) {
                    window.removeEventListener('wheel', window.__hud_wheel_handler);
                    window.removeEventListener('touchmove', window.__hud_touch_handler);
                    window.removeEventListener('keydown', window.__hud_key_handler);
                }

                window.__hud_wheel_handler = preventDefault;
                window.__hud_touch_handler = preventDefault;
                window.__hud_key_handler = preventKeys;

                window.addEventListener('wheel', preventDefault, { passive: false });
                window.addEventListener('touchmove', preventDefault, { passive: false });
                window.addEventListener('keydown', preventKeys, { passive: false });
            })();
        `;
        parentDoc.head.appendChild(script);
    })();
</script>
""", height=1, width=1)
# --- BARRA LATERAL TÁCTICA ---
st.sidebar.image(os.path.join(BASE_DIR, "my_foda_s5.jpeg"), width="stretch")
st.sidebar.markdown("<h3 style='font-family:Orbitron; text-align:center; margin-top: 10px;'>FODA INTELIGENTE</h3>", unsafe_allow_html=True)
st.sidebar.markdown("---")

st.session_state.rol = st.sidebar.selectbox(
    "Perfil del Operador",
    ["Comandante", "Analista", "Observador"],
    index=["Comandante", "Analista", "Observador"].index(st.session_state.rol)
)

st.sidebar.markdown(f"""
<div class="operator-info">
    <div style="margin-bottom: 6px;"><span class="label">Nivel de Acceso:</span> <span class="val">{st.session_state.rol}</span></div>
    <div><span class="label">Gravedad del Sistema:</span> <span class="val gravity-{st.session_state.alert_level.lower()}">{st.session_state.alert_level}</span></div>
</div>
""", unsafe_allow_html=True)

# Recuperación histórica de misiones
st.sidebar.markdown("---")
st.sidebar.markdown("<h5 style='font-family:Orbitron;'>RECOVERY // ARCHIVO DE MISIONES</h5>", unsafe_allow_html=True)
historial = get_all_analyses()
if historial:
    opciones_hist = {f"{r['timestamp'][:19].replace('T', ' ')} {r['perfil']}": r for r in historial}
    seleccion = st.sidebar.selectbox("Seleccione Misión Histórica", list(opciones_hist.keys()))
    if seleccion is not None and st.sidebar.button("RESTAURAR HISTÓRICO", width="stretch"):
        mision_datos = opciones_hist[seleccion]
        raw_data = json.loads(mision_datos['foda_data'])
        # Mapear nombres largos de columnas a iniciales para compatibilidad
        mapping = {
            "Fortalezas": "F", "Debilidades": "D", "Oportunidades": "O", "Amenazas": "A",
            "F": "F", "D": "D", "O": "O", "A": "A"
        }
        mapped_data = {}
        for k, v in raw_data.items():
            mapped_key = mapping.get(k, k)
            mapped_data[mapped_key] = v
        st.session_state.data_foda = mapped_data
        # Actualizar DataFrames en sesión
        for t in ["F", "D", "O", "A"]:
            st.session_state[f"df_{t}"] = sanitizar_dataframe(pd.DataFrame(st.session_state.data_foda.get(t, [])))
        st.session_state.ultimo_resultado = ""  # Forzar recalculación de reporte
        st.session_state.came_estrategias = {
            "A. Estrategia Ofensiva (F + O)": "",
            "B. Estrategia de Supervivencia (D + A)": "",
            "C. Estrategia Defensiva (F + A)": "",
            "D. Estrategia de Reorientación (D + O)": ""
        }
        registrar_evento(f"Restaurada misión histórica ID {mision_datos['id']}")
        st.rerun()
    if st.sidebar.button("BORRAR HISTORIAL", width="stretch"):
        clear_all_analyses()
        registrar_evento("Historial de misiones borrado por completo")
        st.rerun()
else:
    st.sidebar.info("Sin misiones guardadas.")

st.sidebar.markdown("---")
st.sidebar.markdown("<h5 style='font-family:Orbitron;'>DOCUMENTACIÓN</h5>", unsafe_allow_html=True)
if st.session_state.get("ver_manual", False):
    if st.sidebar.button("REGRESAR AL ANÁLISIS", width="stretch"):
        st.session_state.ver_manual = False
        st.rerun()
else:
    if st.sidebar.button("MANUAL DEL OPERADOR", width="stretch"):
        st.session_state.ver_manual = True
        st.rerun()
# =============================================================================
# ESTRUCTURACIÓN DE PESTAÑAS (TABS EQUIDISTANTES Y PRECISAS)
# =============================================================================
# =============================================================================
# PRESENTACIÓN DE PANTALLA PRINCIPAL (MANUAL O PESTAÑAS DE ANÁLISIS)
# =============================================================================
if st.session_state.get("ver_manual", False):
    st.markdown('<h1 class="manual-h1" style="margin-top: 20px;">MANUAL DEL OPERADOR // JOPNAV S-5</h1>', unsafe_allow_html=True)
    
    col_ret1, col_ret2 = st.columns([2, 1])
    with col_ret1:
        st.markdown("<p style='font-size: 15px; color: #88a0b0; font-family: monospace; margin-top: 5px;'>DESCRIPCIÓN GENERAL DEL PROGRAMA FODA INTELIGENTE</p>", unsafe_allow_html=True)
    with col_ret2:
        if st.button("← VOLVER A LA HERRAMIENTA", key="volver_arriba", width="stretch"):
            st.session_state.ver_manual = False
            st.rerun()
        
    st.markdown("""<div class="manual-container">
Este manual proporciona una guía detallada y exhaustiva, redactada para comprender todas las funciones, controles y dinámicas del FODA INTELIGENTE. Diseñado como una herramienta de análisis de apoyo exclusivo a la Sección de Doctrina y Planes Estratégicos (S-5), este sistema integra matrices de diagnóstico, modelado matemático de riesgo y un motor de inteligencia artificial local fuera de línea para preservar la confidencialidad.

<h2 class="manual-h2">1. PANEL DE CONTROL LATERAL (CONTROL DE ACCESO Y ARCHIVO)</h2>
<div class="manual-intro">La barra lateral izquierda actúa como el centro de administración y seguridad del sistema. Desde aquí se definen las autorizaciones del operador y se administran los registros históricos.</div>
<div class="manual-item"><strong>1. Perfil del Operador (Roles)</strong>: Determina el nivel de privilegios funcionales:</div>
<div class="manual-subitem"><strong>1.1. Comandante</strong>: Habilita la formulación, asignación y marcado de directivas operacionales en el Plan de Acción.</div>
<div class="manual-subitem"><strong>1.2. Analista</strong>: Autoriza la edición de matrices FODA, la participación en votaciones estratégicas y la generación de reportes automáticos de IA.</div>
<div class="manual-subitem"><strong>1.3. Observador</strong>: Modo de visualización pasiva. Restringe la interacción, inhabilitando la edición, adición/eliminación de factores, la votación CAME y la creación de directivas.</div>
<div class="manual-item"><strong>2. Archivo de Misiones (Recovery)</strong>: Se conecta de forma persistente a una base de datos local SQLite. Permite consultar análisis previos indexados por fecha, hora y rol. Al presionar <code>RESTAURAR HISTÓRICO</code>, las matrices activas se sobrescriben con los datos históricos guardados, forzando la recalculación en cascada de todo el tablero.</div>
<div class="manual-item"><strong>3. Seguridad Operacional y de Red</strong>:</div>
<div class="manual-subitem"><strong>3.1. Ejecución 100% en Bucle Local</strong>: La aplicación web se ejecuta en la dirección de bucle local (<code>http://localhost:8501</code>). Esto significa que el puerto no está expuesto a la red externa ni a internet a menos que tú lo configures explícitamente en el router.</div>
<div class="manual-subitem"><strong>3.2. Privacidad Absoluta de la IA (Ollama Local)</strong>: A diferencia de otras soluciones que envían tu información a servidores externos (como OpenAI o Claude) a través de Internet siendo vulnerables a intercepciones "Man-in-the-Middle", tu motor de IA corre de forma local en tu máquina (<code>localhost:11434</code>). Ningún dato estratégico sale del equipo físico.</div>
<div class="manual-subitem"><strong>3.3. Prevención de Inyección de Código e Inyección SQL</strong>: En la base de datos SQLite, las consultas son parametrizadas (se usan placeholders), lo que neutraliza por completo cualquier intento de inyección de código SQL malicioso a través de los formularios. Los factores ingresados por el usuario se sanitizan mediante expresiones regulares neutralizando caracteres de control y scripts maliciosos.</div>
<div class="manual-subitem"><strong>3.4. Empaquetado de Seguridad (.exe)</strong>: El empaquetado en un ejecutable cerrado (.exe) dentro de un binario ejecutable comprimido. Esto es un nivel de seguridad: Alto. El usuario solo ve el archivo ejecutable "FODA_S5.exe" siendo imposible leer o modificar el código fuente original.</div>


<h2 class="manual-h2">2. PESTAÑA MATRICES PONDERADAS</h2>
<div class="manual-intro">Es la base de datos de entrada del sistema, dividida en factores internos (Fortalezas y Debilidades) y externos (Oportunidades y Amenazas).</div>
<div class="manual-item"><strong>1. Metodología de ponderación FODA</strong>: Consta de tres fases:</div>
<div class="manual-subitem"><strong>1.1. Asignación de Peso (Importancia)</strong>: A cada factor se le otorga un peso ponderado entre 0.0 y 1.0, dependiendo de su relevancia para la organización, sumando siempre un total de 1.0 (no deberá ser superior a este valor, en cada cuadrante Ej. 1.0 Fortalezas; 1.0 Oportunidades; 1.0 Debilidades; 1.0 Amenazas).</div>
<div class="manual-subitem"><strong>1.2. Calificación (Desempeño)</strong>: Se evalúa la situación actual de cada factor respecto a la empresa. Se utiliza una escala de 1.0 a 5.0, donde 5.0 es una respuesta superior y 1.0 deficiente.</div>
<div class="manual-subitem"><strong>1.3. Cálculo del Total Ponderado</strong>: Se multiplica el peso de cada factor por su calificación. Los factores con los totales más altos son aquellos que requieren atención o inversión prioritaria.</div>
<div class="manual-item"><strong>2. Carga y Mapeo NLP Inteligente</strong>: A través del cargador de archivos, el operador puede subir archivos CSV o Excel (.xlsx). El sistema ejecuta un mapeo semántico NLP avanzado para identificar automáticamente qué columnas corresponden a cada cuadrante del FODA (incluso si tienen nombres informales o abreviados), cargando la información en segundos.</div>
<div class="manual-item"><strong>3. Edición en Tiempo Real sin Latencia</strong>: Las tablas interactivas permiten alterar descripciones de factores, pesos ponderados y calificaciones sin interferencias ni pérdida de foco. Las calificaciones operan en una escala de 1.0 (mínimo) a 5.0 (máximo). Las filas se pueden agregar o eliminar dinámicamente usando los controles interactivos integrados en cada tabla.</div>
<div class="manual-item"><strong>4. Auditoría y Autonormalización MEFI y MEFE</strong>: Las funciones de autonormalización de la Matriz de Evaluación de Factores Internos (MEFI) y la Matriz de Evaluación de Factores Externos (MEFE) tienen el propósito de auditar, corregir y balancear matemáticamente los pesos asignados a los factores de las matrices FODA para que cumplan estrictamente con las reglas metodológicas de planeación estratégica.</div>
<div class="manual-subitem"><strong>4.1. Auditoría de Cumplimiento Metodológico</strong>: El sistema fiscaliza de forma constante que la suma de ponderaciones de cada ámbito de la matriz sume exactamente 1.00, por ejemplo:</div>
<div class="manual-subsubitem"><strong>4.1.1. Ámbito Interno (MEFI)</strong>: Suma de Pesos de Fortalezas (F) + Suma de Pesos de Debilidades (D) = 1.00.</div>
<div class="manual-subsubitem"><strong>4.1.2. Ámbito Externo (MEFE)</strong>: Suma de Pesos de Oportunidades (O) + Suma de Pesos de Amenazas (A) = 1.00.</div>
<div class="manual-subitem"><strong>4.2. Corrección Proporcional</strong>: En caso de desbalance, se activan alertas rojas y los botones de <code>AUTONORMALIZAR MATRIZ</code> que distribuyen proporcionalmente los pesos de manera matemática automática.</div>

<h2 class="manual-h2">3. PESTAÑA DIAGNÓSTICO ESTRATÉGICO</h2>
<div class="manual-intro">Convierte los datos crudos en inteligencia de posicionamiento geográfico/operacional.</div>
<div class="manual-item"><strong>1. Plano de Posicionamiento Vectorial</strong>: Muestra un plano cartesiano interactivo de cuatro cuadrantes donde se grafica el vector resultante (X, Y) derivado de los pesos y calificaciones. Su cuadrante define el enfoque operativo recomendado: Ofensivo (FO), Adaptativo (DO), Defensivo (FA) o de Supervivencia (DA).</div>
<div class="manual-item"><strong>2. Métricas de Balance</strong>: Integra un gráfico de radar que mapea visualmente la distribución de fuerzas de cada factor, facilitando al Estado Mayor la detección de asimetrías o brechas tácticas de forma inmediata.</div>

<h2 class="manual-h2">4. PESTAÑA ANÁLISIS CAUSA-EFECTO</h2>
<div class="manual-intro">El Diagrama de Ishikawa (espina de pescado) se genera 100% de forma dinámica y automática a partir de la información activa que ingresas en las matrices FODA.</div>
<div class="manual-item"><strong>1. La Cabeza del Pescado (El Efecto)</strong>: Se actualiza sola según el Estado Estratégico actual calculado del vector resultante (ej. DOMINIO TOTAL, SUPERVIVENCIA CRÍTICA, etc.).</div>
<div class="manual-item"><strong>2. Las Espinas del Pescado (Las Causas)</strong>:</div>
<div class="manual-subitem"><strong>2.1. Clasificación Automática de Causas</strong>: El sistema analiza el texto de tus Debilidades y Amenazas para distribuirlas en tiempo real dentro del diagrama utilizando palabras clave de nivel estratégico, operacional y táctico:</div>
<div class="manual-subsubitem"><strong>2.1.1. Tecnología</strong>: Clasifica debilidades relacionadas con sistemas, software, ciberdefensa, redes, hardware, infraestructura, servidores, telecomunicaciones o soporte técnico.</div>
<div class="manual-subsubitem"><strong>2.1.2. Personal</strong>: Agrupa debilidades sobre recursos humanos (RRHH), personal, tropa, entrenamiento, liderazgo, reclutamiento, competencias, salarios, fatiga o formación.</div>
<div class="manual-subsubitem"><strong>2.1.3. Procesos</strong>: Reúne de forma predeterminada cualquier otra debilidad operativa interna que no esté catalogada como Tecnología o Personal.</div>
<div class="manual-subsubitem"><strong>2.1.4. Entorno</strong>: Se alimenta de las Amenazas de la matriz externa (factores externos ambientales, geopolíticos, presiones de competidores, clima, etc.), tomando hasta 3 de las incidencias más críticas.</div>

<h2 class="manual-h2">5. PESTAÑA ANÁLISIS DE RIESGO & ESTRÉS</h2>
<div class="manual-intro">Modela la viabilidad de la misión bajo escenarios adversos o simulaciones estocásticas.</div>
<div class="manual-item"><strong>1. Prueba de Estrés</strong>: Deslizadores de hostilidad y pérdida de recursos que alteran los vectores reales simulando situaciones de degradación crítica.</div>
<div class="manual-item"><strong>2. Simulación Monte Carlo (1000 iteraciones)</strong>: Proyecta una nube de dispersión de 1,000 caminos posibles para calcular probabilísticamente el éxito de la misión. Si la probabilidad de éxito disminuye a rangos peligrosos, el semáforo de alerta global cambia dinámicamente.</div>
<div class="manual-item"><strong>3. Máximo Balance Teórico Factible</strong>: Indicador de optimización estratégica derivado de un algoritmo de búsqueda local (búsqueda estocástica de ascenso de colinas) programado en el sistema:</div>
<div class="manual-subitem"><strong>3.1. La Fórmula de Balance (Desempeño)</strong>: El sistema calcula el balance estratégico sumando el balance interno y el externo: Donde cada factor está ponderado por su peso (importancia) y calificación (desempeño).</div>
<div class="manual-subitem"><strong>3.2. ¿Por qué el valor máximo por defecto es 2.0?</strong>: Bajo la configuración estándar de pesos (0.25 por factor), se considera:</div>
<div class="manual-subsubitem"><strong>3.2.1. El "mejor escenario interno"</strong>: Maximizar Fortalezas al límite (5.0) y minimizar Debilidades al mínimo (1.0). Esto da un balance interno máximo de: [(0.25 * 5.0) = 1.25] - [(0.25 * 1.0) = 0.25] = 1.25 - 0.25 = 1.0.</div>
<div class="manual-subsubitem"><strong>3.2.2. El "mejor escenario externo"</strong>: Explotar Oportunidades al límite (5.0) y contener Amenazas al mínimo (1.0). Esto da un balance externo máximo de: [(0.25 * 5.0) = 1.25] - [(0.25 * 1.0) = 0.25] = 1.25 - 0.25 = 1.0.</div>
<div class="manual-subsubitem"><strong>3.2.3. Suma de balances óptimos</strong>: El "máximo teórico" alcanzable bajo esa distribución de pesos (escenario interno y externo) es 1.0 + 1.0 = 2.0.</div>
<div class="manual-subitem"><strong>3.3. La Utilidad de la Métrica</strong>: El número 2.0 (o el puntaje calculado según tus pesos actuales) representa el "techo de rendimiento estratégico" teórico si se lograra mitigar al 100% las debilidades/amenazas y potenciar al máximo las fortalezas/oportunidades.</div>

<h2 class="manual-h2">6. PESTAÑA INFORME IA & EXPORTACIÓN</h2>
<div class="manual-intro">Genera documentación formal clasificada y exportaciones de alta fidelidad.</div>
<div class="manual-item"><strong>1. Motor de Inteligencia Artificial (Ollama + Llama3)</strong>: A partir de las matrices activas, redacta un informe militar clasificado y estructurado (SITUACIÓN GENERAL, AMENAZAS, RIESGOS y LÍNEAS DE ACCIÓN), parafraseado y sin emojis. Los rubros de Amenazas, Riesgos y Líneas de Acción se numeran automáticamente con números arábigos estándar (1., 2., 3., etc.) sin ceros a la izquierda, con sangría alineada y texto justificado. Asimismo, el texto de la "🔵 SITUACIÓN GENERAL" se presenta de forma justificada en prosa.</div>
<div class="manual-item"><strong>2. Exportación en PDF y Excel</strong>: El botón de exportación genera un PDF horizontal optimizado para HUD o un archivo de Excel consolidando los datos en 7 pestañas estructuradas.</div>

<h2 class="manual-h2">7. PESTAÑA MANDO Y BITÁCORA</h2>
<div class="manual-intro">Facilita el control de la reunión táctica en tiempo real:</div>
<div class="manual-item"><strong>1. Mando Compartido (Votación CAME)</strong>: Permite a los analistas emitir votos en favor de las estrategias recomendadas. El consenso y la visualización de porcentajes de decisión se ajustan dinámicamente al total de votantes elegidos en el selector (1 a 25), con barras de progreso de alta estética.</div>
<div class="manual-subitem"><strong>1.1. Fundamentos de la Metodología CAME</strong>: El análisis CAME busca diseñar directivas de acción concretas vinculando cada cuadrante FODA con una línea operativa de respuesta:</div>
<div class="manual-subsubitem"><strong>1.1.1. Corregir (Debilidades)</strong>: Mitigar las debilidades internas de la organización (ámbito interno).</div>
<div class="manual-subsubitem"><strong>1.1.2. Afrontar (Amenazas)</strong>: Definir planes estratégicos para contener riesgos externos de los que no se tiene control directo.</div>
<div class="manual-subsubitem"><strong>1.1.3. Mantener (Fortalezas)</strong>: Robustecer y conservar aquello que ya se realiza de manera sobresaliente.</div>
<div class="manual-subsubitem"><strong>1.1.4. Explotar (Oportunidades)</strong>: Aprovechar los escenarios externos que pueden beneficiar a la organización.</div>
<div class="manual-subitem"><strong>1.2. Tipos de Estrategias CAME Evaluadas</strong>:</div>
<div class="manual-subsubitem"><strong>1.2.1. Estrategia Ofensiva (FO)</strong>: Toma las fortalezas de la organización para aprovechar las oportunidades del entorno. Consiste en enfocarse en aquello en lo que se posee ventaja exclusiva frente a la competencia.</div>
<div class="manual-subsubitem"><strong>1.2.2. Estrategia de Supervivencia (DA)</strong>: Enfrenta debilidades y amenazas buscando mitigar o eliminar vulnerabilidades internas para resistir los riesgos del entorno y asegurar la continuidad de la misión.</div>
<div class="manual-subsubitem"><strong>1.2.3. Estrategia Defensiva (FA)</strong>: Explota las fortalezas acumuladas para contener y neutralizar las amenazas externas, manteniendo la posición de seguridad.</div>
<div class="manual-subsubitem"><strong>1.2.4. Estrategia de Reorientación (DO)</strong>: Detecta fallas internas (debilidades) a corregir para poder aprovechar las oportunidades que de otro modo serían inalcanzables.</div>
<div class="manual-item"><strong>2. Plan de Acción Directivo</strong>: Lista de tareas y directivas de control exclusivo del Comandante.</div>
<div class="manual-subitem"><strong>2.1. Roles tácticos en el Plan de Acción</strong>: El ingreso de directivas y tareas en el plan de acción es exclusivo para el rol de Comandante. Por defecto, cuando abres la herramienta, tu perfil se inicializa como Analista (lo que te permite modificar matrices, votar y generar reportes de IA, pero no formular directivas).</div>
<div class="manual-subitem"><strong>2.2. Cómo habilitar los controles de ingreso de directivas</strong>:</div>
<div class="manual-subsubitem">2.2.1. Ve a la barra lateral izquierda (debajo de la imagen táctica).</div>
<div class="manual-subsubitem">2.2.2. Localiza el selector con el título "Perfil del Operador".</div>
<div class="manual-subsubitem">2.2.3. Cambia tu perfil de Analista a Comandante.</div>
<div class="manual-subsubitem">2.2.4. ¡Listo! Vuelve a la pestaña MANDO Y BITÁCORA y verás que en la sección derecha (ACTION_PLAN) ha aparecido el campo de entrada "Nueva Acción Directiva" y el botón "AÑADIR ACCIÓN".</div>
<div class="manual-item"><strong>3. Bitácora del Turno (Audit Log)</strong>: Registro inmutable de eventos del sistema (modificaciones, votos, normalizaciones) para auditorías de operaciones.</div>

<h2 class="manual-h2">8. ALERTA HUD 100% DINÁMICA (FÓRMULA DE RIESGO REAL)</h2>
<div class="manual-intro">¿Qué función tiene?: Evalúa algorítmicamente en tiempo real la viabilidad de la misión basándose en la probabilidad de éxito de las simulaciones y el nivel de hostilidad de los factores:</div>
<div class="manual-item"><strong>1. VERDE (Normal)</strong>: Probabilidad de éxito > 70% sin amenazas de alto nivel. Indica una postura táctica segura.</div>
<div class="manual-item"><strong>2. AMARILLO (Precaución)</strong>: Probabilidad de éxito entre 40% y 70%. Requiere monitoreo constante y preparación de planes de mitigación.</div>
<div class="manual-item"><strong>3. ROJO (Crítico)</strong>: Probabilidad de éxito < 40%, o si se califica alguna amenaza en la Matriz A con rango severo (>= 4.5), o estrés de simulación crítico (> 20.0). Demanda una acción inmediata de re-evaluación o retirada.</div>
</div>""", unsafe_allow_html=True)
    
    st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
    if st.button("← VOLVER A LA HERRAMIENTA ESTRATÉGICA", key="volver_abajo", width="stretch"):
        st.session_state.ver_manual = False
        st.rerun()
        
    st.markdown("---")
    st.caption("FODA S5 | Comando y Control Estratégico | JOPNAV S-5 | Información Confidencial")
    st.stop()

if "stored_active_tab" not in st.session_state or st.session_state["stored_active_tab"] is None:
    st.session_state["stored_active_tab"] = st.session_state.get("last_valid_tab", "MATRICES PONDERADAS")

active_tab = st.segmented_control(
    "Panel Operativo",
    options=[
        "MATRICES PONDERADAS",
        "DIAGNÓSTICO ESTRATÉGICO",
        "ANÁLISIS CAUSA-EFECTO",
        "FORMULACIÓN CAME",
        "ANÁLISIS DE RIESGO & ESTRÉS",
        "INFORME IA & EXPORTACIÓN",
        "MANDO Y BITÁCORA"
    ],
    key="stored_active_tab",
    label_visibility="collapsed"
)

if active_tab is None:
    st.session_state["stored_active_tab"] = st.session_state.get("last_valid_tab", "MATRICES PONDERADAS")
    st.rerun()
else:
    if st.session_state.get("last_valid_tab") != active_tab:
        reset_pdf_state()
        st.session_state["last_valid_tab"] = active_tab

# -----------------------------------------------------------------------------
# DYNAMIC ROUTING DE PANELES (OPTIMIZACIÓN DE RENDIMIENTO)
# -----------------------------------------------------------------------------
if active_tab == "MATRICES PONDERADAS":
    render_tab_matrices()

elif active_tab == "DIAGNÓSTICO ESTRATÉGICO":
    col_d1, col_d2 = st.columns([3, 2])
    with col_d1:
        st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
        st.markdown('<div class="hud-panel-title">SITREP // PLANO DE POSICIONAMIENTO VECTORIAL</div>', unsafe_allow_html=True)
        st.plotly_chart(grafico_posicionamiento(x, y), width="stretch")
        st.markdown('</div>', unsafe_allow_html=True)
        
    with col_d2:
        st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
        st.markdown('<div class="hud-panel-title">METRICS // COMPORTAMIENTO Y BALANCE</div>', unsafe_allow_html=True)
        st.write(f"**Posición Estratégica:** {estado_actual}")
        alertas(scores)
        st.markdown("---")
        st.plotly_chart(radar_estrategico(scores), width="stretch")
        st.markdown('</div>', unsafe_allow_html=True)
        
    st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
    st.markdown('<div class="hud-panel-title">BALANCE // HISTOGRAMA DE PESOS ACUMULADOS</div>', unsafe_allow_html=True)
    st.plotly_chart(barras_balance(scores), width="stretch")
    st.markdown('</div>', unsafe_allow_html=True)

elif active_tab == "ANÁLISIS CAUSA-EFECTO":
    st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
    st.markdown('<div class="hud-panel-title">ISHIKAVEA // DIAGRAMA DE CAUSA-EFECTO EN TIEMPO REAL</div>', unsafe_allow_html=True)
    st.caption("Diagrama estructurado a partir de las causas internas (Debilidades clasificados según tecnología, procesos y personas) y externas (Amenazas del entorno).")
    
    # Dibujar Diagrama Ishikawa dinámico en base al FODA
    fig_ish = crear_diagrama_ishikawa(df_d, df_a, estado_actual)
    st.plotly_chart(fig_ish, width="stretch")
    st.markdown('</div>', unsafe_allow_html=True)

elif active_tab == "FORMULACIÓN CAME":
    x_val, y_val = scores["interno"], scores["externo"]
    if x_val > 0 and y_val > 0:
        rec_strategy = "A. Estrategia Ofensiva (F + O)"
    elif x_val < 0 and y_val > 0:
        rec_strategy = "D. Estrategia de Reorientación (D + O)"
    elif x_val > 0 and y_val < 0:
        rec_strategy = "C. Estrategia Defensiva (F + A)"
    else:
        rec_strategy = "B. Estrategia de Supervivencia (D + A)"

    came_options = [
        "A. Estrategia Ofensiva (F + O)",
        "B. Estrategia de Supervivencia (D + A)",
        "C. Estrategia Defensiva (F + A)",
        "D. Estrategia de Reorientación (D + O)"
    ]

    if "chosen_came_strategy_val" not in st.session_state:
        st.session_state.chosen_came_strategy_val = rec_strategy

    if "last_rec_strategy" not in st.session_state:
        st.session_state.last_rec_strategy = rec_strategy

    if rec_strategy != st.session_state.last_rec_strategy:
        st.session_state.last_rec_strategy = rec_strategy
        st.session_state.chosen_came_strategy_val = rec_strategy

    col_came1, col_came2 = st.columns(2)
    with col_came1:
        st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
        st.markdown('<div class="hud-panel-title">METODOLOGÍA CAME // CONEXIÓN FODA-CAME</div>', unsafe_allow_html=True)
        st.markdown("""
        <table style="width: 100%; border-collapse: collapse; text-align: left; font-family: monospace; font-size: 14px; color: #e2e8f0;">
            <thead>
                <tr style="border-bottom: 2px solid rgba(0, 242, 255, 0.3); color: #00f2fe;">
                    <th style="padding: 8px; text-align: center;">Factor FODA</th>
                    <th style="padding: 8px; text-align: center;">Origen</th>
                    <th style="padding: 8px; text-align: center;">Acción CAME</th>
                    <th style="padding: 8px; text-align: center;">Enfoque de la Estrategia</th>
                </tr>
            </thead>
            <tbody>
                <tr style="border-bottom: 1px solid rgba(255, 255, 255, 0.05);">
                    <td style="padding: 8px; color: #ffaa00; font-weight: bold;">Debilidades</td>
                    <td style="padding: 8px;">Interno</td>
                    <td style="padding: 8px; color: #ffaa00; font-weight: bold;">Corregir</td>
                    <td style="padding: 8px;">Corregir debilidades internas de la institución y brechas operativas.</td>
                </tr>
                <tr style="border-bottom: 1px solid rgba(255, 255, 255, 0.05);">
                    <td style="padding: 8px; color: #ff0055; font-weight: bold;">Amenazas</td>
                    <td style="padding: 8px;">Externo</td>
                    <td style="padding: 8px; color: #ff0055; font-weight: bold;">Afrontar</td>
                    <td style="padding: 8px;">Afrontar riesgos externos del entorno.</td>
                </tr>
                <tr style="border-bottom: 1px solid rgba(255, 255, 255, 0.05);">
                    <td style="padding: 8px; color: #39FF14; font-weight: bold;">Fortalezas</td>
                    <td style="padding: 8px;">Interno</td>
                    <td style="padding: 8px; color: #39FF14; font-weight: bold;">Mantener</td>
                    <td style="padding: 8px;">Mantener fortalezas y lo que se hace bien.</td>
                </tr>
                <tr>
                    <td style="padding: 8px; color: #00f2fe; font-weight: bold;">Oportunidades</td>
                    <td style="padding: 8px;">Externo</td>
                    <td style="padding: 8px; color: #00f2fe; font-weight: bold;">Explotar</td>
                    <td style="padding: 8px;">Explotar oportunidades externas nacional o internacional.</td>
                </tr>
            </tbody>
        </table>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col_came2:
        st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
        st.markdown('<div class="hud-panel-title">METODOLOGÍA CAME // POSTURAS ESTRATÉGICAS</div>', unsafe_allow_html=True)
        st.markdown("""
        <table style="width: 100%; border-collapse: collapse; text-align: left; font-family: monospace; font-size: 14px; color: #e2e8f0;">
            <thead>
                <tr style="border-bottom: 2px solid rgba(0, 242, 255, 0.3); color: #00f2fe;">
                    <th style="padding: 8px; width: 35%; text-align: center;">Estrategia</th>
                    <th style="padding: 8px; width: 25%; text-align: center;">Objetivo</th>
                    <th style="padding: 8px; width: 40%; text-align: center;">Lógica</th>
                </tr>
            </thead>
            <tbody>
                <tr style="border-bottom: 1px solid rgba(255, 255, 255, 0.05);">
                    <td style="padding: 8px; color: #39FF14; font-weight: bold;">A. Ofensiva (F + O)</td>
                    <td style="padding: 8px; color: #39FF14;">Maximizar fuerzas para capturar beneficios.</td>
                    <td style="padding: 8px;">Enfoque en lo que somos buenos y no los demás para capturar oportunidades inaccesibles.</td>
                </tr>
                <tr style="border-bottom: 1px solid rgba(255, 255, 255, 0.05);">
                    <td style="padding: 8px; color: #ff0055; font-weight: bold;">B. Supervivencia (D + A)</td>
                    <td style="padding: 8px; color: #ff0055;">Minimizar daños y resistir hostilidad.</td>
                    <td style="padding: 8px;">Mitigar/eliminar debilidades internas para dar cara a las amenazas sin desaparecer.</td>
                </tr>
                <tr style="border-bottom: 1px solid rgba(255, 255, 255, 0.05);">
                    <td style="padding: 8px; color: #00f2fe; font-weight: bold;">C. Defensiva (F + A)</td>
                    <td style="padding: 8px; color: #00f2fe;">Explotar fuerzas como escudo.</td>
                    <td style="padding: 8px;">Usar fortalezas para hacer frente a amenazas, manteniendo la posición defensiva.</td>
                </tr>
                <tr>
                    <td style="padding: 8px; color: #ffaa00; font-weight: bold;">D. Reorientación (D + O)</td>
                    <td style="padding: 8px; color: #ffaa00;">Corregir brechas operativas.</td>
                    <td style="padding: 8px;">Corregir fallas internas para poder aprovechar oportunidades que de otra forma serían inalcanzables.</td>
                </tr>
            </tbody>
        </table>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
    st.markdown('<div class="hud-panel-title">FORMULADOR ESTRATÉGICO ASISTIDO POR IA</div>', unsafe_allow_html=True)

    # 1. Panel de Control Superior: Recomendación del Vector y Selector lado a lado
    col_ctrl1, col_ctrl2 = st.columns([1, 1])
    
    with col_ctrl1:
        # Recomendación del Vector
        color_rec = "#39FF14" if "Ofensiva" in rec_strategy else "#00f2fe" if "Defensiva" in rec_strategy else "#ffaa00" if "Reorientación" in rec_strategy else "#ff0055"
        
        # Mapa de colores RGB para efectos glow transparentes
        rgb_map = {
            "#39FF14": "57, 255, 20",
            "#00f2fe": "0, 242, 254",
            "#ffaa00": "255, 170, 0",
            "#ff0055": "255, 0, 85"
        }
        rgb_val = rgb_map.get(color_rec, "0, 242, 254")
        
        st.markdown(f"""
        <div style="background-color: rgba(2, 6, 23, 0.9); border: 1px solid {color_rec}; padding: 12px 18px; border-radius: 4px; box-shadow: inset 0 0 10px rgba(0, 0, 0, 0.5), 0 0 10px rgba({rgb_val}, 0.15); margin-bottom: 10px;">
            <div style="font-family: 'Orbitron', sans-serif; font-size: 11px; text-transform: uppercase; color: #8892b0; letter-spacing: 1.5px;">Recomendación del Vector Táctico</div>
            <div style="font-family: 'Orbitron', sans-serif; font-size: 16px; font-weight: bold; color: {color_rec}; text-shadow: 0 0 8px {color_rec}; margin-top: 4px; letter-spacing: 0.5px;">{rec_strategy}</div>
        </div>
        """, unsafe_allow_html=True)
        
    with col_ctrl2:
        # Selector
        try:
            idx_came = came_options.index(st.session_state.chosen_came_strategy_val)
        except ValueError:
            idx_came = 0
        estrategia_sel = st.selectbox(
            "Seleccione la Postura Estratégica a Formular:",
            options=came_options,
            index=idx_came
        )
        st.session_state.chosen_came_strategy_val = estrategia_sel
        
    if estrategia_sel == "A. Estrategia Ofensiva (F + O)":
        insumos_foda = {"Fortalezas (F)": df_f.to_dict('records'), "Oportunidades (O)": df_o.to_dict('records')}
    elif estrategia_sel == "B. Estrategia de Supervivencia (D + A)":
        insumos_foda = {"Debilidades (D)": df_d.to_dict('records'), "Amenazas (A)": df_a.to_dict('records')}
    elif estrategia_sel == "C. Estrategia Defensiva (F + A)":
        insumos_foda = {"Fortalezas (F)": df_f.to_dict('records'), "Amenazas (A)": df_a.to_dict('records')}
    else:
        insumos_foda = {"Debilidades (D)": df_d.to_dict('records'), "Oportunidades (O)": df_o.to_dict('records')}

    # 2. Panel de Insumos Activos Distribuidos Inteligentemente en 2 Columnas
    st.markdown("""
    <div style="background-color: rgba(2, 6, 23, 0.4); border: 1px solid rgba(0, 242, 255, 0.15); padding: 15px 20px; border-radius: 4px; margin-top: 15px; margin-bottom: 15px;">
        <div style="font-family: 'Orbitron', sans-serif; font-size: 11px; text-transform: uppercase; color: #00f2fe; letter-spacing: 1.5px; border-bottom: 1px solid rgba(0, 242, 255, 0.15); padding-bottom: 6px; margin-bottom: 12px; font-weight: bold; text-shadow: 0 0 5px rgba(0, 242, 255, 0.3);">
            FACTORES CARGADOS PARA EL ANÁLISIS ESTRATÉGICO CAME
        </div>
    """, unsafe_allow_html=True)
    
    col_insumos1, col_insumos2 = st.columns(2)
    categories = list(insumos_foda.keys())
    cat_colors = {
        "Fortalezas (F)": "#39FF14",
        "Oportunidades (O)": "#00f2fe",
        "Debilidades (D)": "#ff0055",
        "Amenazas (A)": "#ff3333"
    }
    
    bg_rgb_map = {
        "#39FF14": "57, 255, 20",
        "#00f2fe": "0, 242, 254",
        "#ff0055": "255, 0, 85",
        "#ff3333": "255, 51, 51"
    }
    
    with col_insumos1:
        cat1 = categories[0]
        items1 = insumos_foda[cat1]
        color1 = cat_colors.get(cat1, "#ffffff")
        bg_rgb1 = bg_rgb_map.get(color1, "0, 242, 254")
        
        st.markdown(f'<div style="font-family: \'Orbitron\', sans-serif; font-size: 13px; font-weight: bold; color: {color1}; margin-bottom: 8px; text-shadow: 0 0 5px {color1};">&raquo; {cat1}</div>', unsafe_allow_html=True)
        if items1:
            for idx, item in enumerate(items1):
                factor_text = item.get('Factor', '')
                peso = item.get('Peso', 0.0)
                calif = item.get('Calificación', 0.0)
                st.markdown(f"""
                <div style="display: flex; justify-content: space-between; align-items: center; background: rgba(255, 255, 255, 0.02); border-left: 3px solid {color1}; padding: 6px 10px; margin-bottom: 4px; border-radius: 0 4px 4px 0; font-family: monospace; font-size: 12px; color: #cbd5e1; border-top: 1px solid rgba(255, 255, 255, 0.01); border-right: 1px solid rgba(255, 255, 255, 0.01);">
                    <span style="flex-grow: 1; padding-right: 10px; text-align: left; font-family: 'Share Tech Mono', monospace; font-size: 13px;">{idx + 1}. {factor_text}</span>
                    <div style="display: flex; gap: 6px; flex-shrink: 0;">
                        <span style="background: rgba(0, 242, 255, 0.08); border: 1px solid rgba(0, 242, 255, 0.25); color: #00f2fe; font-size: 10px; padding: 2px 5px; border-radius: 2px; font-weight: bold; font-family: 'Share Tech Mono', monospace;">Peso: {peso:.2f}</span>
                        <span style="background: rgba({bg_rgb1}, 0.08); border: 1px solid {color1}40; color: {color1}; font-size: 10px; padding: 2px 5px; border-radius: 2px; font-weight: bold; font-family: 'Share Tech Mono', monospace;">Calif: {calif:.1f}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown('<div style="font-family: monospace; font-size: 12px; color: #64748b; font-style: italic; padding: 8px 10px; background: rgba(255,255,255,0.01); border-radius: 4px;">(Ningún factor activo en esta categoría)</div>', unsafe_allow_html=True)

    with col_insumos2:
        cat2 = categories[1]
        items2 = insumos_foda[cat2]
        color2 = cat_colors.get(cat2, "#ffffff")
        bg_rgb2 = bg_rgb_map.get(color2, "0, 242, 254")
        
        st.markdown(f'<div style="font-family: \'Orbitron\', sans-serif; font-size: 13px; font-weight: bold; color: {color2}; margin-bottom: 8px; text-shadow: 0 0 5px {color2};">&raquo; {cat2}</div>', unsafe_allow_html=True)
        if items2:
            for idx, item in enumerate(items2):
                factor_text = item.get('Factor', '')
                peso = item.get('Peso', 0.0)
                calif = item.get('Calificación', 0.0)
                st.markdown(f"""
                <div style="display: flex; justify-content: space-between; align-items: center; background: rgba(255, 255, 255, 0.02); border-left: 3px solid {color2}; padding: 6px 10px; margin-bottom: 4px; border-radius: 0 4px 4px 0; font-family: monospace; font-size: 12px; color: #cbd5e1; border-top: 1px solid rgba(255, 255, 255, 0.01); border-right: 1px solid rgba(255, 255, 255, 0.01);">
                    <span style="flex-grow: 1; padding-right: 10px; text-align: left; font-family: 'Share Tech Mono', monospace; font-size: 13px;">{idx + 1}. {factor_text}</span>
                    <div style="display: flex; gap: 6px; flex-shrink: 0;">
                        <span style="background: rgba(0, 242, 255, 0.08); border: 1px solid rgba(0, 242, 255, 0.25); color: #00f2fe; font-size: 10px; padding: 2px 5px; border-radius: 2px; font-weight: bold; font-family: 'Share Tech Mono', monospace;">Peso: {peso:.2f}</span>
                        <span style="background: rgba({bg_rgb2}, 0.08); border: 1px solid {color2}40; color: {color2}; font-size: 10px; padding: 2px 5px; border-radius: 2px; font-weight: bold; font-family: 'Share Tech Mono', monospace;">Calif: {calif:.1f}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown('<div style="font-family: monospace; font-size: 12px; color: #64748b; font-style: italic; padding: 8px 10px; background: rgba(255,255,255,0.01); border-radius: 4px;">(Ningún factor activo en esta categoría)</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    with st.expander("🛠️ CONFIGURACIÓN DE PROMPT DE IA (CAME)", expanded=False):
        prompt_came_input = st.text_area(
            "Mensaje de Sistema (System Prompt) para CAME. Use {estrategia_tipo}, {objetivo} y {logica} como marcadores de sustitución dinámica.",
            value=st.session_state.prompt_came,
            height=250,
            key="prompt_came_input_ta"
        )
        if prompt_came_input != st.session_state.prompt_came:
            st.session_state.prompt_came = prompt_came_input

    col_btn1, col_btn2, col_btn3 = st.columns([1.5, 1.5, 1])
    generar_came = False
    with col_btn1:
        if st.session_state.rol != "Observador":
            if st.button("EJECUTAR GENERADOR DE ESTRATEGIAS IA", use_container_width=True):
                generar_came = True
        else:
            st.button("EJECUTAR GENERADOR (Bloqueado)", use_container_width=True, disabled=True)
    with col_btn2:
        if st.session_state.rol != "Observador":
            if st.button("GENERAR NUEVA ESTRATEGIA IA", use_container_width=True):
                st.session_state.came_estrategias[estrategia_sel] = ""
                generar_came = True
        else:
            st.button("GENERAR NUEVA ESTRATEGIA IA (Bloqueado)", use_container_width=True, disabled=True)
    with col_btn3:
        if st.session_state.rol != "Observador":
            if st.button("🗑️ LIMPIAR TODO", use_container_width=True):
                st.session_state.came_estrategias = {
                    "A. Estrategia Ofensiva (F + O)": "",
                    "B. Estrategia de Supervivencia (D + A)": "",
                    "C. Estrategia Defensiva (F + A)": "",
                    "D. Estrategia de Reorientación (D + O)": ""
                }
                reset_pdf_state()
                st.toast("Estrategias CAME eliminadas de la memoria.", icon="🗑️")
                st.rerun()
        else:
            st.button("🗑️ LIMPIAR TODO (Bloqueado)", use_container_width=True, disabled=True)

    if st.session_state.came_estrategias[estrategia_sel]:
        st.markdown(f'<div style="font-size: 12px; color: #38bdf8; font-family: monospace; margin-top: 5px; margin-bottom: 10px;">ℹ️ Estrategia ya formulada para la postura seleccionada. Puede volver a generarla si lo desea.</div>', unsafe_allow_html=True)

    came_placeholder = st.empty()
    if st.session_state.came_estrategias[estrategia_sel] and not generar_came:
        with came_placeholder.container():
            st.markdown("##### ESTRATEGIA CAME FORMULADA")
            st.markdown(format_came_estrategia(st.session_state.came_estrategias[estrategia_sel]), unsafe_allow_html=True)

    if generar_came:
        st.session_state.came_estrategias[estrategia_sel] = ""
        came_placeholder.empty()
        
        with came_placeholder.container():
            st.markdown("##### ESTRATEGIA CAME FORMULADA")
            placeholder_came = st.empty()
            
        full_response_came = ""
        try:
            with st.spinner("Conectando con motor de IA local y formulando estrategia..."):
                stream_came = ejecutar_came_ia_stream(estrategia_sel, insumos_foda)
                first_chunk = next(stream_came, None)
                if first_chunk is not None:
                    full_response_came += first_chunk
                    placeholder_came.markdown(format_came_estrategia(full_response_came + " ▒"), unsafe_allow_html=True)
                
            chunk_count = 0
            for chunk in stream_came:
                full_response_came += chunk
                chunk_count += 1
                if chunk_count % 8 == 0:
                    clean_disp = quitar_emojis(full_response_came)
                    placeholder_came.markdown(format_came_estrategia(clean_disp + " ▒"), unsafe_allow_html=True)
            clean_disp = quitar_emojis(full_response_came)
            placeholder_came.markdown(format_came_estrategia(clean_disp), unsafe_allow_html=True)
            st.session_state.came_estrategias[estrategia_sel] = clean_disp
            reset_pdf_state()
            registrar_evento(f"Formulada estrategia CAME: {estrategia_sel}")
            st.rerun()
        except Exception as e:
            st.error(f"Error al formular estrategia CAME con IA: {str(e)}")

    st.markdown('</div>', unsafe_allow_html=True)

elif active_tab == "ANÁLISIS DE RIESGO & ESTRÉS":
    st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
    st.markdown('<div class="hud-panel-title">STRESS // CONTROLES DE PRUEBA DE ESTRÉS OPERATIVO</div>', unsafe_allow_html=True)
    
    col_str1, col_str2 = st.columns(2)
    with col_str1:
        st.slider(
            "Factor de Pérdida de Recursos Internos (Afecta Debilidades)",
            min_value=-30.0, max_value=30.0, value=0.0, step=1.0,
            key="stress_int",
            help="Desplaza los escenarios simulados hacia el cuadrante negativo del eje X (Interno)."
        )
    with col_str2:
        st.slider(
            "Factor de Incremento de Amenazas Externas (Afecta Entorno)",
            min_value=-30.0, max_value=30.0, value=0.0, step=1.0,
            key="stress_ext",
            help="Desplaza los escenarios simulados hacia el cuadrante negativo del eje Y (Externo)."
        )
    st.markdown('</div>', unsafe_allow_html=True)
    
    col_p1, col_p2 = st.columns(2)
    resultados = simulacion_montecarlo(df_f, df_d, df_o, df_a, st.session_state.stress_int, st.session_state.stress_ext)
    prob = probabilidad_exito(resultados)
    
    with col_p1:
        st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
        st.markdown('<div class="hud-panel-title">SIM // DISPERSIÓN DE ESCENARIOS DE MONTE CARLO</div>', unsafe_allow_html=True)
        st.plotly_chart(grafico_simulacion(resultados), width="stretch")
        st.write(f"**Escenarios Exitosos:** {prob}% | **Nivel de Seguridad:** {nivel_riesgo(prob)}")
        st.markdown('</div>', unsafe_allow_html=True)
        
    with col_p2:
        st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
        st.markdown('<div class="hud-panel-title">OMEGA // PROYECCIÓN DE EVOLUCIÓN HISTÓRICA MODIFICADA</div>', unsafe_allow_html=True)
        hist = simular_periodos(df_f, df_d, df_o, df_a, st.session_state.stress_int, st.session_state.stress_ext, T=12)
        st.plotly_chart(grafico_evolucion(hist), width="stretch")
        # Optimización local
        try:
            (mejor_f, mejor_d, mejor_o, mejor_a), score_opt = optimizar(df_f, df_d, df_o, df_a)
            st.write(f"**Máximo balance teórico factible:** {score_opt}")
        except Exception:
            pass
        st.markdown('</div>', unsafe_allow_html=True)

elif active_tab == "INFORME IA & EXPORTACIÓN":
    st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
    st.markdown('<div class="hud-panel-title">CONFIG_IA // PARÁMETROS DEL ANALIZADOR DE IA</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        perfil_opts = ["Sin perfil", "Inteligencia", "Operaciones", "Logística", "Planes"]
        idx_perfil = perfil_opts.index(st.session_state.perfil_val) if st.session_state.perfil_val in perfil_opts else 0
        perfil = st.selectbox("Perfil de Misión", options=perfil_opts, index=idx_perfil)
        if perfil != st.session_state.perfil_val:
            st.session_state.perfil_val = perfil
            limpiar_reporte_ia()
            st.rerun()
    with c2:
        tipo_opts = [
            "Análisis Militar", "Análisis Político", "Análisis Social",
            "Análisis Ciberespacial", "Análisis Económico",
            "Análisis Político/Social/Económico/Militar", "Análisis General"
        ]
        idx_tipo = tipo_opts.index(st.session_state.tipo_ana_val) if st.session_state.tipo_ana_val in tipo_opts else 0
        tipo_ana = st.selectbox("Categoría de Análisis", options=tipo_opts, index=idx_tipo)
        if tipo_ana != st.session_state.tipo_ana_val:
            st.session_state.tipo_ana_val = tipo_ana
            limpiar_reporte_ia()
            st.rerun()
    with c3:
        f_inicio = st.date_input("Inicio de Misión", value=st.session_state.f_inicio_val)
        if f_inicio != st.session_state.f_inicio_val:
            st.session_state.f_inicio_val = f_inicio
            limpiar_reporte_ia()
            st.rerun()
    with c4:
        f_fin = st.date_input("Fin de Misión", value=st.session_state.f_fin_val)
        if f_fin != st.session_state.f_fin_val:
            st.session_state.f_fin_val = f_fin
            limpiar_reporte_ia()
            st.rerun()

    with st.expander("🛠️ CONFIGURACIÓN DE PROMPT DE IA (INFORME FODA)", expanded=False):
        prompt_foda_input = st.text_area(
            "Mensaje de Sistema (System Prompt) para Informe FODA. Use {perfil}, {tipo_ana}, {fecha_i}, {fecha_t}, {x} y {y} como marcadores de sustitución dinámica.",
            value=st.session_state.prompt_foda,
            height=250
        )
        prompt_foda_str = prompt_foda_input or ""
        session_prompt_str = str(st.session_state.prompt_foda or "")
        if prompt_foda_str.replace("\r\n", "\n") != session_prompt_str.replace("\r\n", "\n"):
            st.session_state.prompt_foda = prompt_foda_input
            limpiar_reporte_ia()
            st.rerun()

    col_foda1, col_foda2 = st.columns(2)
    ejecutar = False
    with col_foda1:
        if st.button("EJECUTAR ANÁLISIS ESTRATÉGICO IA", use_container_width=True):
            ejecutar = True
    with col_foda2:
        if st.button("GENERAR NUEVO ANÁLISIS IA", use_container_width=True):
            st.session_state.ultimo_resultado = ""
            ejecutar = True
            
    reporte_placeholder = st.empty()
    if st.session_state.ultimo_resultado and not ejecutar:
        with reporte_placeholder.container():
            st.markdown("##### REPORTE GENERADO EN TIEMPO REAL")
            st.markdown(format_ia_report(st.session_state.ultimo_resultado), unsafe_allow_html=True)

    if ejecutar:
        reset_pdf_state()
        st.session_state.ultimo_resultado = ""
        reporte_placeholder.empty()
        
        with reporte_placeholder.container():
            st.markdown("##### REPORTE GENERADO EN TIEMPO REAL")
            placeholder = st.empty()
            
        foda_data = {
            "Fortalezas": df_f.to_dict('records'),
            "Debilidades": df_d.to_dict('records'),
            "Oportunidades": df_o.to_dict('records'),
            "Amenazas": df_a.to_dict('records')
        }
        full_response = ""
        try:
            with st.spinner("Conectando con motor de IA local y analizando vectores estratégicos..."):
                stream_gen = ejecutar_motor_ia_stream(perfil, tipo_ana, foda_data, f_inicio, f_fin, x, y)
                first_chunk = next(stream_gen, None)
                if first_chunk is not None:
                    full_response += first_chunk
                    disp_text = normalizar_espaciado(sanitizar_emojis_texto_ia(full_response))
                    placeholder.markdown(format_ia_report(disp_text + " ▒"), unsafe_allow_html=True)
                
            chunk_count = 0
            for chunk in stream_gen:
                full_response += chunk
                chunk_count += 1
                if chunk_count % 8 == 0:
                    disp_text = normalizar_espaciado(sanitizar_emojis_texto_ia(full_response))
                    placeholder.markdown(format_ia_report(disp_text + " ▒"), unsafe_allow_html=True)
                
            disp_text = normalizar_espaciado(sanitizar_emojis_texto_ia(full_response))
            placeholder.markdown(format_ia_report(disp_text), unsafe_allow_html=True)
            st.session_state.ultimo_resultado = disp_text
            save_analysis(perfil, tipo_ana, foda_data, x, y)
            registrar_evento(f"Informe de Inteligencia generado: {tipo_ana}")
            
            if "crítico" in full_response.lower() or "alto riesgo" in full_response.lower() or "amenaza" in full_response.lower():
                st.session_state.alert_level = "ROJO"
            else:
                st.session_state.alert_level = "AMARILLO"
            st.rerun()
        except Exception as e:
            st.error(f"Error al conectar con el motor de IA: {str(e)}")
    st.markdown('</div>', unsafe_allow_html=True)

    # Exportador de reportes
    st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
    st.markdown('<div class="hud-panel-title">EXPORT // CONSOLIDACIÓN DE INFORMES</div>', unsafe_allow_html=True)
    if not st.session_state.ultimo_resultado:
        st.warning("ADVERTENCIA: Genere primero un informe de IA antes de exportar.")
    else:
        foda_data = {
            "F": df_f.to_dict('records'),
            "D": df_d.to_dict('records'),
            "O": df_o.to_dict('records'),
            "A": df_a.to_dict('records')
        }
        renderizar_panel_exportacion_informe(
            foda_data=foda_data,
            x=x,
            y=y,
            prob_exito=prob_exito,
            perfil=perfil,
            tipo_ana=tipo_ana,
            f_inicio=f_inicio,
            f_fin=f_fin
        )
    st.markdown('</div>', unsafe_allow_html=True)

elif active_tab == "MANDO Y BITÁCORA":
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
        st.markdown('<div class="hud-panel-title">MANDO COMPARTIDO // VOTACIÓN ESTRATÉGICA CAME</div>', unsafe_allow_html=True)
        
        # Selector de cantidad de votantes y botón de voto al lado
        if st.session_state.rol != "Observador":
            opcion = st.radio(
                "Línea de Acción Recomendada",
                ["FO", "DO", "FA", "DA"],
                horizontal=True,
                key="opcion_vot",
                format_func=lambda x: {
                    "FO": "FO: Ofensiva (F + O)",
                    "DO": "DO: Reorientación (D + O)",
                    "FA": "FA: Defensiva (F + A)",
                    "DA": "DA: Supervivencia (D + A)"
                }[x]
            )
            col_v1, col_v2 = st.columns([1, 1])
            with col_v1:
                total_votantes = st.selectbox(
                    "Cantidad de Votantes",
                    options=list(range(1, 26)),
                    index=st.session_state.get("total_votantes_index", 9), # Por defecto 10
                    key="total_votantes_sel"
                )
                st.session_state.total_votantes = total_votantes
                if isinstance(total_votantes, int):
                    st.session_state.total_votantes_index = list(range(1, 26)).index(total_votantes)
            with col_v2:
                st.markdown("<div style='height: 18px;'></div>", unsafe_allow_html=True) # Espaciador alineación
                if st.button("EMITIR VOTO", width="stretch"):
                    st.session_state.votos[opcion] += 1
                    registrar_evento(f"Voto emitido para estrategia {opcion}")
                    st.rerun()
                    
            if st.button("REINICIAR VOTACIÓN", width="stretch"):
                st.session_state.votos = {"FO": 0, "DO": 0, "FA": 0, "DA": 0}
                registrar_evento("Votación reiniciada")
                st.rerun()
        else:
            total_votantes = st.selectbox(
                "Cantidad de Votantes (Lectura)",
                options=list(range(1, 26)),
                index=st.session_state.get("total_votantes_index", 9),
                key="total_votantes_sel",
                disabled=True
            )
            st.session_state.total_votantes = total_votantes
            
        st.write("Conteo de votos actuales:")
        # Mostrar votos y porcentaje con barras de progreso de alta estética
        opt_names = {
            "FO": "Ofensiva (F + O)",
            "DO": "Reorientación (D + O)",
            "FA": "Defensiva (F + A)",
            "DA": "Supervivencia (D + A)"
        }
        for opt in ["FO", "DO", "FA", "DA"]:
            votos_opt = st.session_state.votos.get(opt, 0)
            pct = (votos_opt / total_votantes) * 100
            st.markdown(f"""
            <div style="margin-bottom: 10px;">
                <div style="display: flex; justify-content: space-between; font-size: 14px; font-family: monospace;">
                    <span>ESTRATEGIA {opt} - {opt_names[opt]}</span>
                    <span>{votos_opt} / {total_votantes} ({pct:.1f}%)</span>
                </div>
                <div style="background-color: rgba(0, 255, 255, 0.1); border: 1px solid rgba(0, 255, 255, 0.3); border-radius: 4px; height: 8px; width: 100%;">
                    <div style="background: linear-gradient(90deg, #00f2fe 0%, #4facfe 100%); width: {min(pct, 100.0):.1f}%; height: 100%; border-radius: 3px;"></div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
        st.info(calcular_consenso(st.session_state.votos, total_votantes))
        st.markdown('</div>', unsafe_allow_html=True)

    with col_c2:
        st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
        st.markdown('<div class="hud-panel-title">ACTION_PLAN // DIRECTIVAS Y SEGUIMIENTO</div>', unsafe_allow_html=True)
        if st.session_state.rol == "Comandante":
            accion = st.text_input("Nueva Acción Directiva", key="nueva_acc")
            if st.button("AÑADIR ACCIÓN"):
                if accion.strip():
                    st.session_state.acciones.append({"accion": accion, "estado": "Pendiente"})
                    reset_pdf_state()
                    registrar_evento(f"Añadida acción al plan: {accion}")
                    st.rerun()
        
        # Listar acciones
        for i, a in enumerate(st.session_state.acciones):
            col_t, col_b = st.columns([4, 1])
            estado = "OK" if a["estado"] == "Completado" else "PENDIENTE"
            col_t.write(f"{estado} {a['accion']}")
            if a["estado"] != "Completado":
                if col_b.button("X", key=f"ok_{i}"):
                    st.session_state.acciones[i]["estado"] = "Completado"
                    reset_pdf_state()
                    registrar_evento(f"Acción completada: {a['accion']}")
                    st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
        
    st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
    st.markdown('<div class="hud-panel-title">AUDIT_LOG // BITÁCORA DEL TURNO ACTIVO</div>', unsafe_allow_html=True)
    for item in reversed(st.session_state.bitacora[-8:]):
        st.write(f"[{item['hora']}] {item['evento']}")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="hud-panel">', unsafe_allow_html=True)
    st.markdown('<div class="hud-panel-title">EXPORT // ACCESO RÁPIDO REPORTE PDF</div>', unsafe_allow_html=True)
    foda_data_mando = {
        "F": df_f.to_dict('records'),
        "D": df_d.to_dict('records'),
        "O": df_o.to_dict('records'),
        "A": df_a.to_dict('records')
    }
    p_val = st.session_state.get("perfil_val", "Sin perfil")
    t_val = st.session_state.get("tipo_ana_val", "Análisis General")
    fi_val = st.session_state.get("f_inicio_val", date.today())
    ff_val = st.session_state.get("f_fin_val", date.today() + timedelta(days=60))
    
    renderizar_panel_exportacion_mando(
        foda_data_mando=foda_data_mando,
        x=x,
        y=y,
        prob_exito=prob_exito,
        p_val=p_val,
        t_val=t_val,
        fi_val=fi_val,
        ff_val=ff_val
    )
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown("---")
st.caption("FODA S5 | Comando y Control Estratégico | JOPNAV S-5 | Información Confidencial")